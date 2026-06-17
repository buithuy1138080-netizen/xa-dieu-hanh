"""
Tạo và seed TẤT CẢ mẫu báo cáo vào database.

Chạy trên VPS:
  docker compose exec backend python /app/seed_report_templates.py
"""
import asyncio
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select

# ── Helpers Excel ─────────────────────────────────────────────────────────────

thin = Side(style="thin")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _wb():
    wb = openpyxl.Workbook()
    ws = wb.active
    return wb, ws

def S(ws, r, c, value="", bold=False, italic=False, size=11, ha="left",
      fill=None, wrap=False, border=False):
    cl = ws.cell(row=r, column=c, value=value)
    cl.font = Font(name="Times New Roman", bold=bold, italic=italic, size=size)
    cl.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fill:
        cl.fill = PatternFill("solid", fgColor=fill)
    if border:
        cl.border = BORDER
    return cl

def merge(ws, r, c1, c2):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

def border_row(ws, r, ncols):
    for c in range(1, ncols+1):
        ws.cell(r, c).border = BORDER


# ══════════════════════════════════════════════════════════════════════════════
# 1. BÁO CÁO NQ57 TUẦN
# ══════════════════════════════════════════════════════════════════════════════
def build_nq57(path: Path):
    wb, ws = _wb()
    ws.title = "BC NQ57"
    for col, w in zip("ABCDEF", [6, 44, 26, 12, 14, 32]):
        ws.column_dimensions[col].width = w
    for r,h in [(1,24),(2,18),(3,18),(4,8),(5,20),(6,20),(7,8),(8,36),(9,5),(10,44),(11,5),(12,8),(13,20),(14,50)]:
        ws.row_dimensions[r].height = h

    S(ws,1,1,"BÁO CÁO TIẾN ĐỘ THEO NGHỊ QUYẾT 57",bold=True,size=14,ha="center"); merge(ws,1,1,6)
    S(ws,2,1,"({{ky_bao_cao}} đến ngày {{den_ngay}})",italic=True,size=11,ha="center"); merge(ws,2,1,6)
    S(ws,3,1,"Đơn vị: {{ten_don_vi}}",size=11); merge(ws,3,1,6)

    S(ws,5,1,"TỔNG QUAN",bold=True,size=11,fill="D9E1F2",border=True); merge(ws,5,1,6)
    for c in range(2,7): ws.cell(5,c).fill=PatternFill("solid",fgColor="D9E1F2"); ws.cell(5,c).border=BORDER
    for c,v,bd in [(1,"Tổng NQ57:",True),(2,"{{tong_nq57}}",False),(3,"Hoàn thành:",True),
                   (4,"{{nq57_hoan_thanh}}",False),(5,"Tiến độ TB:",True),(6,"{{ti_le_nq57}}",False)]:
        S(ws,6,c,v,bold=bd,size=10,border=True)

    for c,h in enumerate(["STT","Nội dung nhiệm vụ","Đơn vị thực hiện","Tiến độ","Hạn","Kết quả / Khó khăn"],1):
        S(ws,8,c,h,bold=True,size=10,ha="center",fill="BDD7EE",wrap=True,border=True)

    ws.cell(9,1,"{{#danh_sach_nq57}}").font = Font(name="Times New Roman",size=7,color="AAAAAA")
    for c,v in enumerate(["{{item.stt}}","{{item.ten}}","{{item.don_vi}}","{{item.tien_do}}","{{item.han}}",""],1):
        S(ws,10,c,v,size=10,ha="center" if c in (1,4,5) else "left",wrap=True,border=True)
    ws.cell(11,1,"{{/danh_sach_nq57}}").font = Font(name="Times New Roman",size=7,color="AAAAAA")

    S(ws,13,1,"Ngày lập: {{ngay_bao_cao}}",italic=True,size=10); merge(ws,13,1,3)
    S(ws,13,5,"THỦ TRƯỞNG ĐƠN VỊ",bold=True,size=10,ha="center"); merge(ws,13,5,6)
    S(ws,14,5,"(Ký, đóng dấu)",italic=True,size=10,ha="center"); merge(ws,14,5,6)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# 2. BÁO CÁO TUẦN TỔNG HỢP
# ══════════════════════════════════════════════════════════════════════════════
def build_weekly(path: Path):
    wb, ws = _wb()
    ws.title = "BC Tuần"
    for col, w in zip("ABCDEF", [6, 38, 20, 14, 14, 20]):
        ws.column_dimensions[col].width = w

    S(ws,1,1,"BÁO CÁO TỔNG HỢP TUẦN",bold=True,size=14,ha="center"); merge(ws,1,1,6)
    S(ws,2,1,"({{ky_bao_cao}} — từ {{tu_ngay}} đến {{den_ngay}})",italic=True,size=11,ha="center"); merge(ws,2,1,6)
    S(ws,3,1,"Đơn vị: {{ten_don_vi}}",size=11); merge(ws,3,1,6)

    # I. Nhiệm vụ
    S(ws,5,1,"I. TÌNH HÌNH THỰC HIỆN NHIỆM VỤ",bold=True,size=11,fill="E2EFDA",border=True); merge(ws,5,1,6)
    for c in range(2,7): ws.cell(5,c).fill=PatternFill("solid",fgColor="E2EFDA"); ws.cell(5,c).border=BORDER
    for r,(a,b,c,d,e,f) in enumerate([
        ("Tổng nhiệm vụ","{{tong_nhiem_vu}}","Hoàn thành","{{nhiem_vu_hoan_thanh}}","Tỉ lệ","{{ti_le_hoan_thanh}}"),
        ("Đang thực hiện","{{nhiem_vu_dang_thuc_hien}}","Chờ xử lý","{{nhiem_vu_cho_xu_ly}}","Quá hạn","{{nhiem_vu_qua_han}}"),
    ], start=6):
        for c,v,bd in [(1,a,True),(2,b,False),(3,c,True),(4,d,False),(5,e,True),(6,f,False)]:
            S(ws,r,c,v,bold=bd,size=10,border=True)

    # II. Văn bản
    S(ws,9,1,"II. VĂN BẢN ĐI / ĐẾN",bold=True,size=11,fill="FFF2CC",border=True); merge(ws,9,1,6)
    for c in range(2,7): ws.cell(9,c).fill=PatternFill("solid",fgColor="FFF2CC"); ws.cell(9,c).border=BORDER
    for c,v,bd in [(1,"Tổng văn bản:",True),(2,"{{tong_van_ban}}",False),(3,"Văn bản đến:",True),
                   (4,"{{van_ban_den}}",False),(5,"Văn bản đi:",True),(6,"{{van_ban_di}}",False)]:
        S(ws,10,c,v,bold=bd,size=10,border=True)

    # III. Nhiệm vụ quá hạn
    S(ws,12,1,"III. NHIỆM VỤ QUÁ HẠN CẦN XỬ LÝ",bold=True,size=11,fill="FCE4D6",border=True); merge(ws,12,1,6)
    for c in range(2,7): ws.cell(12,c).fill=PatternFill("solid",fgColor="FCE4D6"); ws.cell(12,c).border=BORDER
    for c,h in enumerate(["STT","Tên nhiệm vụ","Đơn vị","Hạn","Ngày trễ","Ưu tiên"],1):
        S(ws,13,c,h,bold=True,size=10,ha="center",fill="F4B8A0",border=True)

    ws.cell(14,1,"{{#danh_sach_nhiem_vu_qua_han}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")
    for c,v in enumerate(["{{item.stt}}","{{item.ten}}","{{item.don_vi}}","{{item.han}}","{{item.so_ngay_tre}}","{{item.uu_tien}}"],1):
        S(ws,15,c,v,size=10,ha="center" if c in (1,4,5) else "left",wrap=True,border=True)
    ws.cell(16,1,"{{/danh_sach_nhiem_vu_qua_han}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")

    S(ws,18,1,"Ngày lập: {{ngay_bao_cao}}",italic=True,size=10); merge(ws,18,1,3)
    S(ws,18,5,"THỦ TRƯỞNG ĐƠN VỊ",bold=True,size=10,ha="center"); merge(ws,18,5,6)
    S(ws,19,5,"(Ký, đóng dấu)",italic=True,size=10,ha="center"); merge(ws,19,5,6)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# 3. BÁO CÁO THÁNG TỔNG HỢP
# ══════════════════════════════════════════════════════════════════════════════
def build_monthly(path: Path):
    wb, ws = _wb()
    ws.title = "BC Tháng"
    for col, w in zip("ABCDEFG", [6, 35, 16, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    S(ws,1,1,"BÁO CÁO TỔNG HỢP THÁNG {{thang}}/{{nam}}",bold=True,size=14,ha="center"); merge(ws,1,1,7)
    S(ws,2,1,"(Từ {{tu_ngay}} đến {{den_ngay}})",italic=True,size=11,ha="center"); merge(ws,2,1,7)
    S(ws,3,1,"Đơn vị: {{ten_don_vi}}",size=11); merge(ws,3,1,7)

    # I. Nhiệm vụ
    S(ws,5,1,"I. NHIỆM VỤ",bold=True,size=11,fill="E2EFDA",border=True); merge(ws,5,1,7)
    for c in range(2,8): ws.cell(5,c).fill=PatternFill("solid",fgColor="E2EFDA"); ws.cell(5,c).border=BORDER
    for c,h in enumerate(["","Tổng","Hoàn thành","Đang TH","Chờ XL","Quá hạn","Tỉ lệ HT"],1):
        S(ws,6,c,h,bold=True,size=10,ha="center",fill="C6EFCE",border=True)
    for c,v in enumerate(["Nhiệm vụ","{{tong_nhiem_vu}}","{{nhiem_vu_hoan_thanh}}",
                           "{{nhiem_vu_dang_thuc_hien}}","{{nhiem_vu_cho_xu_ly}}","{{nhiem_vu_qua_han}}","{{ti_le_hoan_thanh}}"],1):
        S(ws,7,c,v,size=10,ha="center" if c>1 else "left",border=True)

    # II. Văn bản
    S(ws,9,1,"II. VĂN BẢN",bold=True,size=11,fill="FFF2CC",border=True); merge(ws,9,1,7)
    for c in range(2,8): ws.cell(9,c).fill=PatternFill("solid",fgColor="FFF2CC"); ws.cell(9,c).border=BORDER
    for c,h in enumerate(["","Tổng","Đã xử lý","Đến","Đi","",""],1):
        S(ws,10,c,h,bold=True,size=10,ha="center",fill="FFEB9C",border=True)
    for c,v in enumerate(["Văn bản","{{tong_van_ban}}","{{van_ban_da_xu_ly}}","{{van_ban_den}}","{{van_ban_di}}","",""],1):
        S(ws,11,c,v,size=10,ha="center" if c>1 else "left",border=True)

    # III. KPI
    S(ws,13,1,"III. CHỈ TIÊU KPI",bold=True,size=11,fill="DDEBF7",border=True); merge(ws,13,1,7)
    for c in range(2,8): ws.cell(13,c).fill=PatternFill("solid",fgColor="DDEBF7"); ws.cell(13,c).border=BORDER
    for c,v,bd in [(1,"Tổng KPI:",True),(2,"{{tong_kpi}}",False),(3,"Tiến độ TB:",True),(4,"{{ti_le_kpi}}",False),(5,"",False),(6,"",False),(7,"",False)]:
        S(ws,14,c,v,bold=bd,size=10,border=True)

    # IV. NQ57
    S(ws,16,1,"IV. NGHỊ QUYẾT 57",bold=True,size=11,fill="FCE4D6",border=True); merge(ws,16,1,7)
    for c in range(2,8): ws.cell(16,c).fill=PatternFill("solid",fgColor="FCE4D6"); ws.cell(16,c).border=BORDER
    for c,v,bd in [(1,"Tổng NQ57:",True),(2,"{{tong_nq57}}",False),(3,"Hoàn thành:",True),(4,"{{nq57_hoan_thanh}}",False),(5,"Tiến độ:",True),(6,"{{ti_le_nq57}}",False),(7,"",False)]:
        S(ws,17,c,v,bold=bd,size=10,border=True)

    # V. Phân tích đơn vị
    S(ws,19,1,"V. PHÂN TÍCH THEO ĐƠN VỊ",bold=True,size=11,fill="EAF0FB",border=True); merge(ws,19,1,7)
    for c in range(2,8): ws.cell(19,c).fill=PatternFill("solid",fgColor="EAF0FB"); ws.cell(19,c).border=BORDER
    for c,h in enumerate(["STT","Đơn vị","","Tổng NV","Hoàn thành","Tỉ lệ",""],1):
        S(ws,20,c,h,bold=True,size=10,ha="center",fill="B8CCE4",border=True)

    ws.cell(21,1,"{{#phan_tich_don_vi}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")
    for c,v in enumerate(["{{item.stt}}","{{item.ten_don_vi}}","","{{item.tong}}","{{item.hoan_thanh}}","{{item.ti_le}}",""],1):
        S(ws,22,c,v,size=10,ha="center" if c in (1,4,5,6) else "left",wrap=True,border=True)
    ws.cell(23,1,"{{/phan_tich_don_vi}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")

    S(ws,25,1,"Ngày lập: {{ngay_bao_cao}}",italic=True,size=10); merge(ws,25,1,3)
    S(ws,25,6,"THỦ TRƯỞNG ĐƠN VỊ",bold=True,size=10,ha="center"); merge(ws,25,6,7)
    S(ws,26,6,"(Ký, đóng dấu)",italic=True,size=10,ha="center"); merge(ws,26,6,7)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# 4. BÁO CÁO CHỈ ĐẠO
# ══════════════════════════════════════════════════════════════════════════════
def build_directive(path: Path):
    wb, ws = _wb()
    ws.title = "BC Chỉ đạo"
    for col, w in zip("ABCDEF", [6, 40, 22, 14, 14, 18]):
        ws.column_dimensions[col].width = w

    S(ws,1,1,"BÁO CÁO THỰC HIỆN CHỈ ĐẠO",bold=True,size=14,ha="center"); merge(ws,1,1,6)
    S(ws,2,1,"({{ky_bao_cao}} — từ {{tu_ngay}} đến {{den_ngay}})",italic=True,size=11,ha="center"); merge(ws,2,1,6)
    S(ws,3,1,"Đơn vị: {{ten_don_vi}}",size=11); merge(ws,3,1,6)

    # Tổng quan
    S(ws,5,1,"TỔNG QUAN",bold=True,size=11,fill="D9E1F2",border=True); merge(ws,5,1,6)
    for c in range(2,7): ws.cell(5,c).fill=PatternFill("solid",fgColor="D9E1F2"); ws.cell(5,c).border=BORDER
    for c,v,bd in [(1,"Tổng chỉ đạo:",True),(2,"{{tong_chi_dao}}",False),(3,"Nhiệm vụ liên quan:",True),
                   (4,"{{tong_nhiem_vu}}",False),(5,"Hoàn thành:",True),(6,"{{nhiem_vu_hoan_thanh}}",False)]:
        S(ws,6,c,v,bold=bd,size=10,border=True)

    # Bảng nhiệm vụ quá hạn theo chỉ đạo
    S(ws,8,1,"NHIỆM VỤ QUÁ HẠN",bold=True,size=11,fill="FCE4D6",border=True); merge(ws,8,1,6)
    for c in range(2,7): ws.cell(8,c).fill=PatternFill("solid",fgColor="FCE4D6"); ws.cell(8,c).border=BORDER
    for c,h in enumerate(["STT","Tên nhiệm vụ","Đơn vị thực hiện","Hạn","Ngày trễ","Ưu tiên"],1):
        S(ws,9,c,h,bold=True,size=10,ha="center",fill="F4B8A0",border=True)

    ws.cell(10,1,"{{#danh_sach_nhiem_vu_qua_han}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")
    for c,v in enumerate(["{{item.stt}}","{{item.ten}}","{{item.don_vi}}","{{item.han}}","{{item.so_ngay_tre}}","{{item.uu_tien}}"],1):
        S(ws,11,c,v,size=10,ha="center" if c in (1,4,5) else "left",wrap=True,border=True)
    ws.cell(12,1,"{{/danh_sach_nhiem_vu_qua_han}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")

    S(ws,14,1,"Ngày lập: {{ngay_bao_cao}}",italic=True,size=10); merge(ws,14,1,3)
    S(ws,14,5,"THỦ TRƯỞNG ĐƠN VỊ",bold=True,size=10,ha="center"); merge(ws,14,5,6)
    S(ws,15,5,"(Ký, đóng dấu)",italic=True,size=10,ha="center"); merge(ws,15,5,6)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# 5. BÁO CÁO KPI
# ══════════════════════════════════════════════════════════════════════════════
def build_kpi(path: Path):
    wb, ws = _wb()
    ws.title = "BC KPI"
    for col, w in zip("ABCDEF", [6, 40, 18, 14, 14, 14]):
        ws.column_dimensions[col].width = w

    S(ws,1,1,"BÁO CÁO TIẾN ĐỘ CHỈ TIÊU KPI",bold=True,size=14,ha="center"); merge(ws,1,1,6)
    S(ws,2,1,"({{ky_bao_cao}} — Năm {{nam}})",italic=True,size=11,ha="center"); merge(ws,2,1,6)
    S(ws,3,1,"Đơn vị: {{ten_don_vi}}",size=11); merge(ws,3,1,6)

    # Tổng quan
    S(ws,5,1,"TỔNG QUAN KPI",bold=True,size=11,fill="DDEBF7",border=True); merge(ws,5,1,6)
    for c in range(2,7): ws.cell(5,c).fill=PatternFill("solid",fgColor="DDEBF7"); ws.cell(5,c).border=BORDER
    for c,v,bd in [(1,"Tổng chỉ tiêu KPI:",True),(2,"{{tong_kpi}}",False),
                   (3,"Tiến độ trung bình:",True),(4,"{{ti_le_kpi}}",False),(5,"",False),(6,"",False)]:
        S(ws,6,c,v,bold=bd,size=10,border=True)

    # KPI NQ57
    S(ws,8,1,"CHỈ TIÊU NQ57",bold=True,size=11,fill="FCE4D6",border=True); merge(ws,8,1,6)
    for c in range(2,7): ws.cell(8,c).fill=PatternFill("solid",fgColor="FCE4D6"); ws.cell(8,c).border=BORDER
    for c,v,bd in [(1,"Tổng NQ57:",True),(2,"{{tong_nq57}}",False),(3,"Hoàn thành:",True),
                   (4,"{{nq57_hoan_thanh}}",False),(5,"Tiến độ:",True),(6,"{{ti_le_nq57}}",False)]:
        S(ws,9,c,v,bold=bd,size=10,border=True)

    # Phân tích theo đơn vị
    S(ws,11,1,"PHÂN TÍCH NHIỆM VỤ THEO ĐƠN VỊ",bold=True,size=11,fill="EAF0FB",border=True); merge(ws,11,1,6)
    for c in range(2,7): ws.cell(11,c).fill=PatternFill("solid",fgColor="EAF0FB"); ws.cell(11,c).border=BORDER
    for c,h in enumerate(["STT","Đơn vị","","Tổng NV","Hoàn thành","Tỉ lệ"],1):
        S(ws,12,c,h,bold=True,size=10,ha="center",fill="B8CCE4",border=True)

    ws.cell(13,1,"{{#phan_tich_don_vi}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")
    for c,v in enumerate(["{{item.stt}}","{{item.ten_don_vi}}","","{{item.tong}}","{{item.hoan_thanh}}","{{item.ti_le}}"],1):
        S(ws,14,c,v,size=10,ha="center" if c in (1,4,5,6) else "left",wrap=True,border=True)
    ws.cell(15,1,"{{/phan_tich_don_vi}}").font=Font(name="Times New Roman",size=7,color="AAAAAA")

    S(ws,17,1,"Ngày lập: {{ngay_bao_cao}}",italic=True,size=10); merge(ws,17,1,3)
    S(ws,17,5,"THỦ TRƯỞNG ĐƠN VỊ",bold=True,size=10,ha="center"); merge(ws,17,5,6)
    S(ws,18,5,"(Ký, đóng dấu)",italic=True,size=10,ha="center"); merge(ws,18,5,6)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# 6. MẪU TỰ CHỈNH (tổng hợp tất cả biến)
# ══════════════════════════════════════════════════════════════════════════════
def build_custom(path: Path):
    wb, ws = _wb()
    ws.title = "Mẫu tùy chỉnh"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 50

    S(ws,1,1,"MẪU BÁO CÁO TÙY CHỈNH",bold=True,size=14,ha="center"); merge(ws,1,1,2)
    S(ws,2,1,"({{ky_bao_cao}} — {{tu_ngay}} đến {{den_ngay}})",italic=True,size=11,ha="center"); merge(ws,2,1,2)
    S(ws,3,1,"Đơn vị: {{ten_don_vi}}",size=11); merge(ws,3,1,2)
    S(ws,4,1,"Ngày lập: {{ngay_bao_cao}}",italic=True,size=10); merge(ws,4,1,2)

    S(ws,6,1,"NỘI DUNG BÁO CÁO (tùy chỉnh theo nhu cầu)",bold=True,size=11); merge(ws,6,1,2)

    # Guide sheet
    ws2 = wb.create_sheet("📌 Tất cả biến")
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 50
    S(ws2,1,1,"TẤT CẢ BIẾN CÓ THỂ DÙNG TRONG MẪU",bold=True,size=12); merge(ws2,1,1,2)

    rows = [
        ("── BIẾN ĐƠN (scalar) ──",""),
        ("{{ten_don_vi}}","Tên đơn vị"),
        ("{{ngay_bao_cao}}","Ngày lập báo cáo (hôm nay)"),
        ("{{ky_bao_cao}}","Nhãn kỳ báo cáo"),
        ("{{tu_ngay}}","Từ ngày"),
        ("{{den_ngay}}","Đến ngày"),
        ("{{thang}}","Tháng số"),
        ("{{quy}}","Quý (I/II/III/IV)"),
        ("{{nam}}","Năm"),
        ("{{tong_nhiem_vu}}","Tổng nhiệm vụ"),
        ("{{nhiem_vu_hoan_thanh}}","Nhiệm vụ hoàn thành"),
        ("{{nhiem_vu_dang_thuc_hien}}","Đang thực hiện"),
        ("{{nhiem_vu_cho_xu_ly}}","Chờ xử lý"),
        ("{{nhiem_vu_qua_han}}","Quá hạn"),
        ("{{ti_le_hoan_thanh}}","Tỉ lệ hoàn thành (%)"),
        ("{{tong_van_ban}}","Tổng văn bản"),
        ("{{van_ban_da_xu_ly}}","Đã xử lý"),
        ("{{van_ban_den}}","Văn bản đến"),
        ("{{van_ban_di}}","Văn bản đi"),
        ("{{tong_kpi}}","Tổng KPI"),
        ("{{ti_le_kpi}}","Tiến độ KPI (%)"),
        ("{{tong_nq57}}","Tổng NQ57"),
        ("{{nq57_hoan_thanh}}","NQ57 hoàn thành"),
        ("{{ti_le_nq57}}","Tiến độ NQ57 (%)"),
        ("{{tong_chi_dao}}","Tổng chỉ đạo"),
        ("",""),
        ("── VÒNG LẶP: {{#danh_sach_nq57}} ──","fields: stt, ma, ten, nhom, don_vi, tien_do, trang_thai, han"),
        ("── VÒNG LẶP: {{#danh_sach_nhiem_vu_qua_han}} ──","fields: stt, ten, han, don_vi, uu_tien, so_ngay_tre"),
        ("── VÒNG LẶP: {{#phan_tich_don_vi}} ──","fields: stt, ten_don_vi, tong, hoan_thanh, ti_le"),
    ]
    for i,(a,b) in enumerate(rows, 2):
        ws2.cell(i,1,a).font = Font(name="Times New Roman",bold=a.startswith("──"),size=10)
        ws2.cell(i,2,b).font = Font(name="Times New Roman",size=10)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


# ══════════════════════════════════════════════════════════════════════════════
# SEED VÀO DATABASE
# ══════════════════════════════════════════════════════════════════════════════

TEMPLATES = [
    {
        "builder": build_nq57,
        "name": "Báo cáo tiến độ NQ57 tuần",
        "category": "nq57",
        "description": "Báo cáo tiến độ thực hiện Nghị quyết 57 hàng tuần — tự động điền số liệu",
        "file": "mau_nq57.xlsx",
    },
    {
        "builder": build_weekly,
        "name": "Báo cáo tuần tổng hợp",
        "category": "weekly",
        "description": "Tổng hợp nhiệm vụ, văn bản, danh sách quá hạn theo tuần",
        "file": "mau_tuan.xlsx",
    },
    {
        "builder": build_monthly,
        "name": "Báo cáo tháng tổng hợp",
        "category": "monthly",
        "description": "Tổng hợp toàn diện: nhiệm vụ, văn bản, KPI, NQ57, phân tích đơn vị",
        "file": "mau_thang.xlsx",
    },
    {
        "builder": build_directive,
        "name": "Báo cáo thực hiện chỉ đạo",
        "category": "directive",
        "description": "Tình hình thực hiện chỉ đạo, nhiệm vụ quá hạn cần xử lý",
        "file": "mau_chi_dao.xlsx",
    },
    {
        "builder": build_kpi,
        "name": "Báo cáo tiến độ KPI",
        "category": "kpi",
        "description": "Tiến độ chỉ tiêu KPI, NQ57 và phân tích theo đơn vị",
        "file": "mau_kpi.xlsx",
    },
    {
        "builder": build_custom,
        "name": "Mẫu tùy chỉnh (tất cả biến)",
        "category": "custom",
        "description": "Mẫu cơ sở có sheet hướng dẫn đầy đủ tất cả biến — dùng để tạo báo cáo riêng",
        "file": "mau_tuy_chinh.xlsx",
    },
]


async def seed_all():
    from app.core.database import AsyncSessionLocal
    from app.core.config import settings
    from app.models.report_template import ReportTemplate
    from app.models.user import User
    from app.services import template_engine

    async with AsyncSessionLocal() as db:
        admin = (await db.execute(
            select(User).where(User.role == "admin").limit(1)
        )).scalar_one_or_none()
        if not admin:
            admin = (await db.execute(select(User).limit(1))).scalar_one_or_none()
        if not admin:
            print("⚠️  Không tìm thấy user, dừng"); return

        print(f"  Dùng user: {admin.username} (id={admin.id})\n")

        for tpl_def in TEMPLATES:
            name = tpl_def["name"]
            category = tpl_def["category"]
            print(f"── {name} [{category}]")

            existing = (await db.execute(
                select(ReportTemplate).where(
                    ReportTemplate.name == name,
                    ReportTemplate.category == category,
                )
            )).scalar_one_or_none()
            if existing:
                print(f"   ⚠️  Đã tồn tại (id={existing.id}), bỏ qua\n"); continue

            # Tạo file Excel
            tmp = Path(f"/tmp/{tpl_def['file']}")
            tpl_def["builder"](tmp)
            print(f"   File: {tmp} ({tmp.stat().st_size:,} bytes)")

            # Copy vào thư mục templates
            dest_dir = Path(settings.UPLOAD_DIR) / "templates" / category
            dest_dir.mkdir(parents=True, exist_ok=True)
            safe = re.sub(r"[^\w\-.]", "_", name)
            dest = dest_dir / f"{safe}_v1.xlsx"
            dest.write_bytes(tmp.read_bytes())

            scalars, lists = await asyncio.to_thread(
                template_engine.parse_variables, str(dest)
            )

            rec = ReportTemplate(
                name=name,
                category=category,
                description=tpl_def["description"],
                file_ext="xlsx",
                file_path=str(dest),
                file_size=dest.stat().st_size,
                variables_json=scalars,
                list_variables_json=lists,
                version=1,
                is_active=True,
                created_by=admin.id,
            )
            db.add(rec)
            await db.commit()
            await db.refresh(rec)
            print(f"   ✅ id={rec.id}  scalar={scalars}  list={lists}\n")

    print("═══ HOÀN THÀNH ═══")


asyncio.run(seed_all())
