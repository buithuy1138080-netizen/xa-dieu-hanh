"""
excel_import.py — shared helpers for parsing and generating Excel import templates.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Column definitions ────────────────────────────────────────────────────────

TASK_COLUMNS = [
    {"key": "title",         "header": "Tiêu đề nhiệm vụ (*)",        "width": 40, "required": True},
    {"key": "description",   "header": "Mô tả",                        "width": 30, "required": False},
    {"key": "responsible_unit", "header": "Đơn vị phụ trách",          "width": 25, "required": False},
    {"key": "due_date",      "header": "Deadline (dd/mm/yyyy)",        "width": 18, "required": False},
    {"key": "priority",      "header": "Ưu tiên (low/medium/high/urgent)", "width": 28, "required": False},
    {"key": "start_date",    "header": "Ngày bắt đầu (dd/mm/yyyy)",    "width": 22, "required": False},
]

NQ57_COLUMNS = [
    {"key": "code",          "header": "Mã nhiệm vụ",                  "width": 14, "required": False},
    {"key": "title",         "header": "Tên nhiệm vụ (*)",             "width": 45, "required": True},
    {"key": "group",         "header": "Nhóm/Lĩnh vực",               "width": 20, "required": False},
    {"key": "target",        "header": "Chỉ tiêu cần đạt",            "width": 30, "required": False},
    {"key": "responsible_unit", "header": "Đơn vị phụ trách",         "width": 25, "required": False},
    {"key": "deadline",      "header": "Deadline (dd/mm/yyyy)",        "width": 18, "required": False},
    {"key": "progress",      "header": "Tiến độ (%)",                  "width": 13, "required": False},
    {"key": "status",        "header": "Trạng thái (pending/in_progress/completed/delayed)", "width": 42, "required": False},
]

KPI_COLUMNS = [
    {"key": "code",          "header": "Mã chỉ tiêu",                  "width": 14, "required": False},
    {"key": "title",         "header": "Tên chỉ tiêu (*)",             "width": 40, "required": True},
    {"key": "category",      "header": "Danh mục (Kinh tế/Xã hội/...)", "width": 28, "required": False},
    {"key": "unit",          "header": "Đơn vị tính (%/người/...)",    "width": 22, "required": False},
    {"key": "target_value",  "header": "Mục tiêu (*)",                 "width": 13, "required": True},
    {"key": "current_value", "header": "Giá trị hiện tại",            "width": 18, "required": False},
    {"key": "year",          "header": "Năm (*)",                      "width": 10, "required": True},
    {"key": "period",        "header": "Kỳ (monthly/quarterly/yearly)", "width": 30, "required": False},
    {"key": "responsible_unit", "header": "Đơn vị phụ trách",         "width": 25, "required": False},
    {"key": "deadline",      "header": "Deadline (dd/mm/yyyy)",        "width": 18, "required": False},
]


# ── Template generator ────────────────────────────────────────────────────────

def _build_template(columns: list[dict], sample_rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1E40AF")
    req_fill    = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    sample_font = Font(color="475569", italic=True, size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 36

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col["header"])
        cell.font = header_font
        cell.fill = req_fill if col.get("required") else header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    for row_idx, row_data in enumerate(sample_rows, start=2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = sample_font
            cell.alignment = left

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def task_template() -> bytes:
    sample = [
        ["Hoàn thiện báo cáo tháng 5", "Tổng hợp kết quả công tác", "Phòng Hành chính", "31/05/2026", "high", "01/05/2026"],
        ["Triển khai phần mềm quản lý", "", "Phòng Kỹ thuật", "30/06/2026", "medium", ""],
    ]
    return _build_template(TASK_COLUMNS, sample)


def nq57_template() -> bytes:
    sample = [
        ["NQ57-001", "Xây dựng hạ tầng mạng cáp quang toàn xã", "Hạ tầng số", "Phủ sóng 100% hộ dân", "Phòng Kỹ thuật", "31/12/2026", "45", "in_progress"],
        ["NQ57-002", "Triển khai dịch vụ công trực tuyến mức độ 4", "Chính phủ số", "80% dịch vụ đạt mức 4", "Phòng Hành chính", "30/09/2026", "20", "pending"],
    ]
    return _build_template(NQ57_COLUMNS, sample)


def kpi_template() -> bytes:
    sample = [
        ["KPI-001", "Tỷ lệ hộ nghèo giảm", "Xã hội", "%", "5.0", "7.2", "2026", "yearly", "Phòng LĐ-TB&XH", "31/12/2026"],
        ["KPI-002", "Thu ngân sách xã", "Kinh tế", "Tỷ đồng", "12.5", "8.3", "2026", "yearly", "Phòng Tài chính", "31/12/2026"],
    ]
    return _build_template(KPI_COLUMNS, sample)


# ── Row parsers ───────────────────────────────────────────────────────────────

def _parse_date(val: Any) -> str | None:
    if not val:
        return None
    if isinstance(val, (datetime,)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return None


def _parse_float(val: Any) -> float:
    try:
        return float(str(val).strip().replace(",", "."))
    except (ValueError, TypeError):
        return 0.0


def _parse_int(val: Any, default: int = 0) -> int:
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _str(val: Any) -> str:
    return str(val).strip() if val is not None else ""


def parse_tasks(data: bytes) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records, errors = [], []
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        title = _str(row[0] if len(row) > 0 else "")
        if not title:
            errors.append(f"Hàng {i}: 'Tiêu đề' là bắt buộc")
            continue
        records.append({
            "title":            title,
            "description":      _str(row[1] if len(row) > 1 else ""),
            "responsible_unit": _str(row[2] if len(row) > 2 else ""),
            "due_date_str":     _parse_date(row[3] if len(row) > 3 else None),
            "priority":         _str(row[4] if len(row) > 4 else "").lower() or "medium",
            "start_date_str":   _parse_date(row[5] if len(row) > 5 else None),
        })
    return records, errors


def parse_nq57(data: bytes) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records, errors = [], []
    valid_statuses = {"pending", "in_progress", "completed", "delayed"}
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        title = _str(row[1] if len(row) > 1 else "")
        if not title:
            errors.append(f"Hàng {i}: 'Tên nhiệm vụ' là bắt buộc")
            continue
        status = _str(row[7] if len(row) > 7 else "").lower()
        if status not in valid_statuses:
            status = "pending"
        progress = min(100, max(0, _parse_int(row[6] if len(row) > 6 else 0)))
        records.append({
            "code":             _str(row[0] if len(row) > 0 else "") or None,
            "title":            title,
            "group":            _str(row[2] if len(row) > 2 else "") or None,
            "target":           _str(row[3] if len(row) > 3 else "") or None,
            "responsible_unit": _str(row[4] if len(row) > 4 else "") or None,
            "deadline_str":     _parse_date(row[5] if len(row) > 5 else None),
            "progress":         progress,
            "status":           status,
        })
    return records, errors


def parse_kpi(data: bytes) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records, errors = [], []
    valid_periods = {"monthly", "quarterly", "yearly"}
    from datetime import date as dt_date
    current_year = dt_date.today().year

    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        title = _str(row[1] if len(row) > 1 else "")
        if not title:
            errors.append(f"Hàng {i}: 'Tên chỉ tiêu' là bắt buộc")
            continue
        target = _parse_float(row[4] if len(row) > 4 else 0)
        if target == 0:
            errors.append(f"Hàng {i}: 'Mục tiêu' phải lớn hơn 0")
            continue
        year = _parse_int(row[6] if len(row) > 6 else current_year, default=current_year)
        period = _str(row[7] if len(row) > 7 else "").lower()
        if period not in valid_periods:
            period = "yearly"
        records.append({
            "code":             _str(row[0] if len(row) > 0 else "") or None,
            "title":            title,
            "category":         _str(row[2] if len(row) > 2 else "") or None,
            "unit":             _str(row[3] if len(row) > 3 else "") or None,
            "target_value":     target,
            "current_value":    _parse_float(row[5] if len(row) > 5 else 0),
            "year":             year,
            "period":           period,
            "responsible_unit": _str(row[8] if len(row) > 8 else "") or None,
            "deadline_str":     _parse_date(row[9] if len(row) > 9 else None),
        })
    return records, errors
