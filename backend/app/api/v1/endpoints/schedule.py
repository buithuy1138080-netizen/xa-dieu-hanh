"""Module Lịch Công Tác — API endpoints."""
from __future__ import annotations

import io
import logging
from datetime import date, datetime, time, timedelta, timezone
from math import ceil
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from pydantic import BaseModel, ConfigDict
from sqlalchemy import and_, delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.deps import get_current_user, get_db
from app.models.schedule import ScheduleItem, ScheduleReminder
from app.models.staff import Staff
from app.models.user import User
from app.models.zalo import ZaloUserLink

logger = logging.getLogger(__name__)
router = APIRouter()

VALID_SESSIONS = {"sang", "chieu", "ca_ngay", "toi"}
VALID_REMIND_MINUTES = {15, 30, 60, 120, 1440}  # 1440 = 1 ngày trước

# Chỉ hiển thị 5 lãnh đạo này trong module lịch công tác
SCHEDULE_LEADER_NAMES = [
    "Nguyễn Duy Hòa",
    "Bùi Thị Lý",
    "Bùi Minh Hải",
    "Nguyễn Tài Nghệ",
    "Phạm Thị Non",
]


# ── Schemas ───────────────────────────────────────────────────────────────────

class LeaderMin(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    full_name: str
    position: str | None = None
    employee_code: str | None = None


class ScheduleItemRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    leader_id: int
    leader: LeaderMin | None = None
    title: str
    location: str | None = None
    note: str | None = None
    work_date: date
    session: str
    start_time: time | None = None
    zalo_remind: bool
    remind_before_minutes: int
    created_by: int | None = None
    created_at: datetime
    updated_at: datetime | None = None


class ScheduleItemCreate(BaseModel):
    leader_id: int
    title: str
    location: str | None = None
    note: str | None = None
    work_date: date
    session: str = "sang"
    start_time: str | None = None       # "08:00" hoặc None
    zalo_remind: bool = False
    remind_before_minutes: int = 30


class ScheduleItemUpdate(BaseModel):
    leader_id: int | None = None
    title: str | None = None
    location: str | None = None
    note: str | None = None
    work_date: date | None = None
    session: str | None = None
    start_time: str | None = None
    zalo_remind: bool | None = None
    remind_before_minutes: int | None = None


class ReminderLogRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    schedule_id: int
    leader_id: int
    zalo_user_id: str | None = None
    scheduled_at: datetime
    sent_at: datetime | None = None
    status: str
    error_msg: str | None = None
    retry_count: int
    created_at: datetime


class PaginatedSchedule(BaseModel):
    items: list[ScheduleItemRead]
    total: int
    page: int
    size: int
    pages: int


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_time(t_str: str | None) -> time | None:
    if not t_str:
        return None
    try:
        parts = t_str.strip().split(":")
        return time(int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)
    except Exception:
        return None


def _calc_remind_dt(work_date: date, start_time: time | None, before_minutes: int) -> datetime | None:
    """Tính thời điểm cần gửi nhắc."""
    if start_time is None:
        # Mặc định: sáng 8h hoặc trước 1440 phút (ngày hôm trước 8h)
        base_dt = datetime.combine(work_date, time(8, 0)).replace(tzinfo=timezone.utc)
    else:
        base_dt = datetime.combine(work_date, start_time).replace(tzinfo=timezone.utc)
    remind_dt = base_dt - timedelta(minutes=before_minutes)
    # Không gửi nhắc trong quá khứ
    if remind_dt <= datetime.now(timezone.utc):
        return None
    return remind_dt


async def _upsert_reminder(db: AsyncSession, item: ScheduleItem) -> None:
    """Tạo hoặc cập nhật reminder record cho 1 lịch."""
    # Xóa pending cũ
    await db.execute(
        delete(ScheduleReminder).where(
            ScheduleReminder.schedule_id == item.id,
            ScheduleReminder.status == "pending",
        )
    )

    if not item.zalo_remind:
        return

    remind_dt = _calc_remind_dt(item.work_date, item.start_time, item.remind_before_minutes)
    if remind_dt is None:
        return

    # Tìm zalo_user_id của lãnh đạo qua Staff → user_id → ZaloUserLink
    staff = await db.get(Staff, item.leader_id)
    zalo_uid: str | None = None
    phone: str | None = None
    if staff and staff.user_id:
        link = (await db.execute(
            select(ZaloUserLink).where(
                ZaloUserLink.user_id == staff.user_id,
                ZaloUserLink.is_active == True,
            )
        )).scalar_one_or_none()
        if link:
            zalo_uid = link.zalo_user_id
            phone = link.zalo_phone

    reminder = ScheduleReminder(
        schedule_id=item.id,
        leader_id=item.leader_id,
        zalo_user_id=zalo_uid,
        recipient_phone=phone,
        scheduled_at=remind_dt,
        status="pending",
    )
    db.add(reminder)


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("", response_model=PaginatedSchedule)
async def list_schedule(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    leader_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    session: str | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = (
        select(ScheduleItem)
        .options(selectinload(ScheduleItem.leader))
        .where(ScheduleItem.deleted_at.is_(None))
    )
    if leader_id:
        q = q.where(ScheduleItem.leader_id == leader_id)
    if date_from:
        q = q.where(ScheduleItem.work_date >= date_from)
    if date_to:
        q = q.where(ScheduleItem.work_date <= date_to)
    if session:
        q = q.where(ScheduleItem.session == session)

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar_one()
    items = (await db.execute(
        q.order_by(ScheduleItem.work_date.asc(), ScheduleItem.start_time.asc().nulls_last())
        .offset((page - 1) * size).limit(size)
    )).scalars().all()

    return PaginatedSchedule(
        items=[ScheduleItemRead.model_validate(i) for i in items],
        total=total, page=page, size=size,
        pages=max(1, ceil(total / size)),
    )


@router.get("/week")
async def week_view(
    week_start: date = Query(..., description="Ngày đầu tuần (Thứ 2), format: YYYY-MM-DD"),
    leader_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Lịch tuần: trả về dict keyed by leader_id → list items theo ngày."""
    week_end = week_start + timedelta(days=6)

    q = (
        select(ScheduleItem)
        .options(selectinload(ScheduleItem.leader))
        .where(
            ScheduleItem.deleted_at.is_(None),
            ScheduleItem.work_date >= week_start,
            ScheduleItem.work_date <= week_end,
        )
        .order_by(ScheduleItem.work_date, ScheduleItem.start_time.asc().nulls_last())
    )
    if leader_id:
        q = q.where(ScheduleItem.leader_id == leader_id)

    items = (await db.execute(q)).scalars().all()

    # Lấy danh sách lãnh đạo
    leaders = (await db.execute(
        select(Staff).where(
            Staff.is_active == True,
            Staff.full_name.in_(SCHEDULE_LEADER_NAMES),
        ).order_by(Staff.full_name)
    )).scalars().all()

    # Cấu trúc dữ liệu theo lãnh đạo × ngày
    result = []
    days = [week_start + timedelta(days=i) for i in range(7)]

    for leader in leaders:
        leader_items = [i for i in items if i.leader_id == leader.id]
        by_day = {}
        for d in days:
            by_day[d.isoformat()] = [
                ScheduleItemRead.model_validate(i)
                for i in leader_items if i.work_date == d
            ]
        result.append({
            "leader": LeaderMin.model_validate(leader),
            "days": by_day,
        })

    return {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "days": [d.isoformat() for d in days],
        "leaders": result,
    }


@router.get("/leaders")
async def list_leaders(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Danh sách lãnh đạo dùng cho dropdown."""
    items = (await db.execute(
        select(Staff).where(
            Staff.is_active == True,
            Staff.full_name.in_(SCHEDULE_LEADER_NAMES),
        ).order_by(Staff.full_name)
    )).scalars().all()
    return [LeaderMin.model_validate(i) for i in items]


@router.get("/reminders/logs", response_model=list[ReminderLogRead])
async def reminder_logs(
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
    leader_id: int | None = None,
    limit: int = Query(100, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    q = (
        select(ScheduleReminder)
        .join(ScheduleItem, ScheduleReminder.schedule_id == ScheduleItem.id)
        .order_by(ScheduleReminder.scheduled_at.desc())
    )
    if status:
        q = q.where(ScheduleReminder.status == status)
    if leader_id:
        q = q.where(ScheduleReminder.leader_id == leader_id)
    if date_from:
        q = q.where(ScheduleItem.work_date >= date_from)
    if date_to:
        q = q.where(ScheduleItem.work_date <= date_to)

    rows = (await db.execute(q.limit(limit))).scalars().all()
    return [ReminderLogRead.model_validate(r) for r in rows]


@router.get("/export")
async def export_excel(
    week_start: date = Query(...),
    leader_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Xuất Excel lịch tuần."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl chưa được cài đặt")

    week_end = week_start + timedelta(days=6)
    days = [week_start + timedelta(i) for i in range(7)]
    day_names = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
    vn_months = ["tháng 1","tháng 2","tháng 3","tháng 4","tháng 5","tháng 6","tháng 7","tháng 8","tháng 9","tháng 10","tháng 11","tháng 12"]

    q = (
        select(ScheduleItem)
        .options(selectinload(ScheduleItem.leader))
        .where(
            ScheduleItem.deleted_at.is_(None),
            ScheduleItem.work_date.between(week_start, week_end),
        )
        .order_by(ScheduleItem.work_date, ScheduleItem.start_time.asc().nulls_last())
    )
    if leader_id:
        q = q.where(ScheduleItem.leader_id == leader_id)

    items = (await db.execute(q)).scalars().all()

    leaders = (await db.execute(
        select(Staff).where(
            Staff.is_active == True,
            Staff.full_name.in_(SCHEDULE_LEADER_NAMES),
        ).order_by(Staff.full_name)
    )).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Lịch tuần {week_start.strftime('%d.%m')}"

    RED = "C0392B"; WHITE = "FFFFFF"; LIGHT_RED = "FADBD8"
    header_fill = PatternFill("solid", fgColor=RED)
    sub_fill    = PatternFill("solid", fgColor=LIGHT_RED)
    header_font = Font(bold=True, color=WHITE, name="Times New Roman", size=11)
    body_font   = Font(name="Times New Roman", size=10)
    center_al   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al     = Alignment(vertical="top", wrap_text=True)
    thin_side   = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "LỊCH CÔNG TÁC TUẦN"
    ws["A1"].font = Font(bold=True, size=13, color=RED, name="Times New Roman")
    ws["A1"].alignment = center_al

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Từ ngày {week_start.strftime('%d/%m/%Y')} đến {week_end.strftime('%d/%m/%Y')}"
    ws["A2"].alignment = center_al
    ws["A2"].font = Font(size=11, name="Times New Roman")

    # Header row
    row = 3
    headers = ["LÃNH ĐẠO"] + [f"{dn}\n{d.strftime('%d/%m/%Y')}" for dn, d in zip(day_names, days)]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_al
        cell.border = thin_border

    ws.column_dimensions["A"].width = 25
    for ci in range(2, 10):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    # Data rows
    for leader in leaders:
        leader_items = [i for i in items if i.leader_id == leader.id]
        row += 1
        ws.cell(row=row, column=1, value=f"{leader.full_name}\n{leader.position or ''}").alignment = center_al
        ws.cell(row=row, column=1).font = Font(bold=True, name="Times New Roman", size=10)
        ws.cell(row=row, column=1).border = thin_border

        for di, d in enumerate(days, 2):
            day_items = [i for i in leader_items if i.work_date == d]
            txt = ""
            for it in day_items:
                t = it.start_time.strftime("%H:%M") if it.start_time else ""
                session_lbl = {"sang": "Sáng", "chieu": "Chiều", "ca_ngay": "Cả ngày", "toi": "Tối"}.get(it.session, it.session)
                line = f"{session_lbl}"
                if t:
                    line += f" {t}"
                line += f": {it.title}"
                if it.location:
                    line += f"\nĐịa điểm: {it.location}"
                txt += line + "\n"
            cell = ws.cell(row=row, column=di, value=txt.strip())
            cell.alignment = left_al
            cell.font = body_font
            cell.border = thin_border

        ws.row_dimensions[row].height = max(40, 15 * max(1, max((len(leader_items) for d in days if any(i.work_date == d for i in leader_items)), default=1)))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"lich_tuan_{week_start.strftime('%d%m%Y')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("", response_model=ScheduleItemRead, status_code=201)
async def create_schedule(
    body: ScheduleItemCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if body.session not in VALID_SESSIONS:
        raise HTTPException(400, f"session phải là: {VALID_SESSIONS}")

    item = ScheduleItem(
        leader_id=body.leader_id,
        title=body.title.strip(),
        location=(body.location or "").strip() or None,
        note=(body.note or "").strip() or None,
        work_date=body.work_date,
        session=body.session,
        start_time=_parse_time(body.start_time),
        zalo_remind=body.zalo_remind,
        remind_before_minutes=body.remind_before_minutes,
        created_by=current_user.id,
    )
    db.add(item)
    try:
        await db.flush()
    except Exception as exc:
        await db.rollback()
        logger.error("schedule create flush error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Lỗi lưu lịch: {exc}")

    await _upsert_reminder(db, item)
    new_id = item.id
    await db.commit()

    result = (await db.execute(
        select(ScheduleItem)
        .options(selectinload(ScheduleItem.leader))
        .where(ScheduleItem.id == new_id)
    )).scalar_one()
    return ScheduleItemRead.model_validate(result)


@router.get("/{item_id}", response_model=ScheduleItemRead)
async def get_schedule(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    item = (await db.execute(
        select(ScheduleItem)
        .options(selectinload(ScheduleItem.leader))
        .where(ScheduleItem.id == item_id, ScheduleItem.deleted_at.is_(None))
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Không tìm thấy lịch")
    return ScheduleItemRead.model_validate(item)


@router.put("/{item_id}", response_model=ScheduleItemRead)
async def update_schedule(
    item_id: int,
    body: ScheduleItemUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    item = (await db.execute(
        select(ScheduleItem).where(ScheduleItem.id == item_id, ScheduleItem.deleted_at.is_(None))
    )).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "Không tìm thấy lịch")

    changed = False
    for field in ["leader_id", "title", "location", "note", "work_date", "session",
                  "zalo_remind", "remind_before_minutes"]:
        val = getattr(body, field)
        if val is not None:
            setattr(item, field, val)
            changed = True
    if body.start_time is not None:
        item.start_time = _parse_time(body.start_time)
        changed = True

    if changed:
        item.updated_at = datetime.now(timezone.utc)
        await _upsert_reminder(db, item)
        await db.commit()

    item_id = item.id
    result = (await db.execute(
        select(ScheduleItem)
        .options(selectinload(ScheduleItem.leader))
        .where(ScheduleItem.id == item_id)
    )).scalar_one()
    return ScheduleItemRead.model_validate(result)


@router.delete("/{item_id}", status_code=204)
async def delete_schedule(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    item = await db.get(ScheduleItem, item_id)
    if not item or item.deleted_at:
        raise HTTPException(404, "Không tìm thấy lịch")

    item.deleted_at = datetime.now(timezone.utc)
    # Hủy các reminder pending
    await db.execute(
        delete(ScheduleReminder).where(
            ScheduleReminder.schedule_id == item_id,
            ScheduleReminder.status == "pending",
        )
    )
    await db.commit()


# ── Sao chép lịch sang lãnh đạo khác ─────────────────────────────────────────

class CopyScheduleRequest(BaseModel):
    item_id: int
    leader_ids: list[int]          # danh sách lãnh đạo đích
    work_date: Optional[date] = None  # nếu None: giữ nguyên ngày gốc
    zalo_remind: Optional[bool] = None  # nếu None: giữ nguyên cài đặt gốc


@router.post("/copy", response_model=list[ScheduleItemRead], status_code=201)
async def copy_schedule(
    body: CopyScheduleRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Sao chép 1 lịch sang nhiều lãnh đạo khác (có thể đổi ngày)."""
    # Lấy lịch nguồn
    source = (await db.execute(
        select(ScheduleItem).where(
            ScheduleItem.id == body.item_id,
            ScheduleItem.deleted_at.is_(None),
        )
    )).scalar_one_or_none()
    if not source:
        raise HTTPException(404, "Không tìm thấy lịch nguồn")

    if not body.leader_ids:
        raise HTTPException(400, "Cần chọn ít nhất 1 lãnh đạo")

    target_date    = body.work_date or source.work_date
    target_remind  = body.zalo_remind if body.zalo_remind is not None else source.zalo_remind

    created = []
    for leader_id in body.leader_ids:
        # Bỏ qua nếu trùng lãnh đạo + cùng ngày (tránh duplicate)
        existing = (await db.execute(
            select(ScheduleItem).where(
                ScheduleItem.leader_id == leader_id,
                ScheduleItem.work_date == target_date,
                ScheduleItem.session == source.session,
                ScheduleItem.title == source.title,
                ScheduleItem.deleted_at.is_(None),
            )
        )).scalar_one_or_none()
        if existing:
            continue

        new_item = ScheduleItem(
            leader_id=leader_id,
            title=source.title,
            location=source.location,
            note=source.note,
            work_date=target_date,
            session=source.session,
            start_time=source.start_time,
            zalo_remind=target_remind,
            remind_before_minutes=source.remind_before_minutes,
            created_by=current_user.id,
        )
        db.add(new_item)
        await db.flush()
        await _upsert_reminder(db, new_item)
        created.append(new_item)

    created_ids = [item.id for item in created]
    await db.commit()

    result = []
    for cid in created_ids:
        r = (await db.execute(
            select(ScheduleItem)
            .options(selectinload(ScheduleItem.leader))
            .where(ScheduleItem.id == cid)
        )).scalar_one()
        result.append(ScheduleItemRead.model_validate(r))

    return result
