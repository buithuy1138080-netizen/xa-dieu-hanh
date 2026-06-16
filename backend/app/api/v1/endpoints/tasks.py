from __future__ import annotations

import shutil
import uuid
from datetime import date, datetime, timezone
from math import ceil
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, ConfigDict, model_validator
from sqlalchemy import and_, case, delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.staff import Staff
from app.models.task import Task, TaskAuditLog, TaskAttachment, TaskComment, TaskDepartment
from app.models.user import User

router = APIRouter()


async def _get_user_dept_id(db: AsyncSession, user_id: int) -> int | None:
    """Get the department_id of the user via their Staff record."""
    staff = (await db.execute(select(Staff).where(Staff.user_id == user_id))).scalar_one_or_none()
    return staff.department_id if staff else None


async def _check_task_write_permission(
    db: AsyncSession,
    current_user: User,
    lead_dept_id: int | None,
    action: str,
    task_id: int | None = None,
) -> None:
    """Enforce role-based write permissions.

    admin / leader : full access
    manager        : create/edit for own dept (lead or coordinating); NO delete
    staff          : create own tasks; edit own tasks (creator/assignee); NO delete
    """
    if current_user.role in ("admin", "leader"):
        return

    user_dept = await _get_user_dept_id(db, current_user.id)

    if current_user.role == "staff":
        # Xóa: không cho phép
        if action == "delete":
            raise HTTPException(403, "Nhân viên không có quyền xóa nhiệm vụ")
        # Tạo mới: cho phép (nhân viên tạo nhiệm vụ cho chính mình)
        if action == "create":
            return
        # Cập nhật: chỉ được sửa nhiệm vụ do mình tạo hoặc được giao cho mình
        if action == "update" and task_id:
            t = (await db.execute(
                select(Task).where(Task.id == task_id, Task.deleted_at.is_(None))
            )).scalar_one_or_none()
            if t is None:
                raise HTTPException(404, "Không tìm thấy nhiệm vụ")
            is_creator  = t.created_by == current_user.id
            is_assignee = t.assignee_id == current_user.id
            # Kiểm tra giao qua bảng Staff
            if not is_creator and not is_assignee and t.assignee_staff_id:
                staff_rec = (await db.execute(
                    select(Staff).where(Staff.user_id == current_user.id)
                )).scalar_one_or_none()
                is_assignee = staff_rec is not None and t.assignee_staff_id == staff_rec.id
            if not is_creator:
                raise HTTPException(403, "Nhân viên chỉ được cập nhật nhiệm vụ do chính mình tạo ra")
        return

    if current_user.role == "manager":
        # Manager không được xóa nhiệm vụ
        if action == "delete":
            raise HTTPException(403, "Quản lý không có quyền xóa nhiệm vụ")

        # Kiểm tra đơn vị phụ trách hoặc phối hợp
        is_lead = lead_dept_id and lead_dept_id == user_dept
        is_coord = False
        if task_id and user_dept:
            coord = (await db.execute(
                select(TaskDepartment).where(
                    TaskDepartment.task_id == task_id,
                    TaskDepartment.department_id == user_dept,
                )
            )).scalar_one_or_none()
            is_coord = coord is not None

        if not is_lead and not is_coord:
            raise HTTPException(403, "Quản lý chỉ được thao tác nhiệm vụ của đơn vị mình phụ trách hoặc phối hợp")


from app.core.config import settings as _settings
UPLOAD_DIR = Path(_settings.UPLOAD_DIR) / "tasks"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

VALID_STATUSES   = {"pending", "in_progress", "completed", "cancelled", "overdue"}
VALID_PRIORITIES = {"low", "medium", "high", "urgent"}
ALLOWED_UPLOAD_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".png", ".jpg", ".jpeg"}
VALID_SORT_COLS  = {"created_at", "updated_at", "due_date", "title", "priority", "status", "progress_percent"}


# ── Schemas ───────────────────────────────────────────────────────────────────

class DeptMin(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    name: str
    short_name: str | None = None
    code: str | None = None


class UserMin(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    full_name: str | None = None
    username: str


class StaffMin(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    full_name: str
    position: str | None = None
    employee_code: str | None = None
    department_id: int | None = None


class DocMin(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    title: str
    doc_number: str | None = None


class DirectiveMin(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    title: str


class TaskDeptRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    department_id: int
    role: str
    department: DeptMin | None = None


class TaskCommentRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    task_id: int
    user_id: int
    content: str
    created_at: datetime
    user: UserMin | None = None


class TaskAttachmentRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    task_id: int
    user_id: int
    filename: str
    file_path: str
    file_size: int
    created_at: datetime


class TaskAuditRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    action: str
    field: str | None = None
    old_value: str | None = None
    new_value: str | None = None
    created_at: datetime
    user: UserMin | None = None


class TaskRead(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: int
    task_code: str | None = None
    title: str
    description: str | None = None
    content_summary: str | None = None
    status: str
    priority: str
    progress_percent: int
    is_project: bool = False
    project_type: str | None = None
    budget_amount: float | None = None
    budget_disbursed: float | None = None
    start_date: date | None = None
    due_date: datetime | None = None
    completed_at: datetime | None = None
    incoming_document_id: int | None = None
    outgoing_document_id: int | None = None
    directive_id: int | None = None
    program_id: int | None = None
    parent_task_id: int | None = None
    task_type: str = "regular"
    task_group: str | None = None
    created_by: int
    updated_by: int | None = None
    assignee_id: int | None = None
    assignee_staff_id: int | None = None
    supervising_user_id: int | None = None
    lead_department_id: int | None = None
    reminder_enabled: bool
    overdue_warning: bool
    completion_note: str | None = None
    deleted_at: datetime | None = None
    created_at: datetime
    updated_at: datetime | None = None
    is_overdue: bool = False
    project_ids: list[int] = []
    subtasks_count: int = 0
    creator: UserMin | None = None
    assignee: UserMin | None = None
    assignee_staff: StaffMin | None = None
    lead_department: DeptMin | None = None


class TaskReadDetail(TaskRead):
    updater: UserMin | None = None
    supervisor: UserMin | None = None
    incoming_document: DocMin | None = None
    outgoing_document: DocMin | None = None
    directive: DirectiveMin | None = None
    departments: list[TaskDeptRead] = []
    comments: list[TaskCommentRead] = []
    attachments: list[TaskAttachmentRead] = []
    audit_logs: list[TaskAuditRead] = []
    subtasks: list["TaskRead"] = []


class TaskCreate(BaseModel):
    title: str
    description: str | None = None
    content_summary: str | None = None
    priority: str = "medium"
    is_project: bool = False
    project_type: str | None = None
    budget_amount: float | None = None
    budget_disbursed: float | None = None
    start_date: date | None = None
    due_date: datetime | None = None
    program_id: int | None = None
    parent_task_id: int | None = None
    incoming_document_id: int | None = None
    outgoing_document_id: int | None = None
    directive_id: int | None = None
    assignee_id: int | None = None
    assignee_staff_id: int | None = None
    supervising_user_id: int | None = None
    lead_department_id: int | None = None
    coordinating_department_ids: list[int] = []
    reminder_enabled: bool = False

    @model_validator(mode='after')
    def check_dates(self):
        if self.start_date and self.due_date:
            due = self.due_date.date() if isinstance(self.due_date, datetime) else self.due_date
            if due < self.start_date:
                raise ValueError('due_date phải lớn hơn hoặc bằng start_date')
        return self


class TaskUpdate(BaseModel):
    title: str | None = None
    description: str | None = None
    content_summary: str | None = None
    priority: str | None = None
    is_project: bool | None = None
    project_type: str | None = None
    budget_amount: float | None = None
    budget_disbursed: float | None = None
    start_date: date | None = None
    due_date: datetime | None = None
    program_id: int | None = None
    parent_task_id: int | None = None
    incoming_document_id: int | None = None
    outgoing_document_id: int | None = None
    directive_id: int | None = None
    assignee_id: int | None = None
    assignee_staff_id: int | None = None
    supervising_user_id: int | None = None
    lead_department_id: int | None = None
    coordinating_department_ids: list[int] | None = None
    reminder_enabled: bool | None = None
    completion_note: str | None = None

    @model_validator(mode='after')
    def check_dates(self):
        if self.start_date and self.due_date:
            due = self.due_date.date() if isinstance(self.due_date, datetime) else self.due_date
            if due < self.start_date:
                raise ValueError('due_date phải lớn hơn hoặc bằng start_date')
        return self


class TaskStatusUpdate(BaseModel):
    status: str
    completion_note: str | None = None


class TaskProgressUpdate(BaseModel):
    progress_percent: int
    completion_note: str | None = None


class TaskCommentCreate(BaseModel):
    content: str


class TaskDeptAdd(BaseModel):
    department_id: int
    role: str = "coordinating"


class TaskStats(BaseModel):
    total: int = 0
    pending: int = 0
    in_progress: int = 0
    completed: int = 0
    cancelled: int = 0
    overdue: int = 0
    high_priority: int = 0
    urgent_priority: int = 0
    avg_progress: float = 0.0


class PaginatedTasks(BaseModel):
    items: list[TaskRead]
    total: int
    page: int
    page_size: int
    pages: int


# ── Helpers ───────────────────────────────────────────────────────────────────

async def _next_task_code(db: AsyncSession) -> str:
    """Generate next task code using MAX(id substring) to avoid race conditions."""
    year = datetime.now().year
    prefix = f"NV-{year}-"
    result = await db.execute(
        select(Task.task_code)
        .where(Task.task_code.like(f"{prefix}%"))
        .order_by(Task.task_code.desc())
        .limit(1)
    )
    last_code = result.scalar_one_or_none()
    if last_code:
        try:
            last_seq = int(last_code.split("-")[-1])
        except (ValueError, IndexError):
            last_seq = 0
        next_seq = last_seq + 1
    else:
        next_seq = 1
    return f"{prefix}{next_seq:04d}"


def _is_overdue(t: Task) -> bool:
    if t.status in ("completed", "cancelled", "overdue"):
        return False
    if t.due_date is None:
        return False
    now = datetime.now(timezone.utc)
    due = t.due_date if t.due_date.tzinfo else t.due_date.replace(tzinfo=timezone.utc)
    return due < now


def _to_read(t: Task) -> dict:
    d = {c.key: getattr(t, c.key) for c in t.__mapper__.column_attrs}
    d["is_overdue"] = _is_overdue(t)
    d["subtasks_count"] = len(t.subtasks) if hasattr(t, "subtasks") and t.subtasks is not None else 0
    if hasattr(t, "creator") and t.creator is not None:
        d["creator"] = {"id": t.creator.id, "full_name": t.creator.full_name, "username": t.creator.username}
    else:
        d["creator"] = None
    if hasattr(t, "assignee") and t.assignee is not None:
        d["assignee"] = {"id": t.assignee.id, "full_name": t.assignee.full_name, "username": t.assignee.username}
    else:
        d["assignee"] = None
    if hasattr(t, "assignee_staff") and t.assignee_staff is not None:
        s = t.assignee_staff
        d["assignee_staff"] = {
            "id": s.id, "full_name": s.full_name,
            "position": s.position, "employee_code": s.employee_code,
            "department_id": s.department_id,
        }
    else:
        d["assignee_staff"] = None
    if hasattr(t, "lead_department") and t.lead_department is not None:
        ld = t.lead_department
        d["lead_department"] = {"id": ld.id, "name": ld.name, "short_name": getattr(ld, "short_name", None), "code": getattr(ld, "code", None)}
    else:
        d["lead_department"] = None
    return d


_DETAIL_LOADS = [
    selectinload(Task.creator),
    selectinload(Task.updater),
    selectinload(Task.assignee),
    selectinload(Task.assignee_staff),
    selectinload(Task.supervisor),
    selectinload(Task.lead_department),
    selectinload(Task.incoming_document),
    selectinload(Task.outgoing_document),
    selectinload(Task.directive),
    selectinload(Task.departments).selectinload(TaskDepartment.department),
    selectinload(Task.comments).selectinload(TaskComment.user),
    selectinload(Task.attachments),
    selectinload(Task.audit_logs).selectinload(TaskAuditLog.user),
    selectinload(Task.subtasks).selectinload(Task.creator),
    selectinload(Task.subtasks).selectinload(Task.assignee),
    selectinload(Task.subtasks).selectinload(Task.lead_department),
]

_LIST_LOADS = [
    selectinload(Task.creator),
    selectinload(Task.assignee),
    selectinload(Task.assignee_staff),
    selectinload(Task.lead_department),
    selectinload(Task.subtasks),
]


async def _get_task(db: AsyncSession, task_id: int, detail: bool = False) -> Task:
    loads = _DETAIL_LOADS if detail else _LIST_LOADS
    q = select(Task).where(Task.id == task_id, Task.deleted_at.is_(None))
    for ld in loads:
        q = q.options(ld)
    result = await db.execute(q)
    t = result.scalar_one_or_none()
    if t is None:
        raise HTTPException(status_code=404, detail="Task not found")
    return t


async def _audit(
    db: AsyncSession, task_id: int, user_id: int,
    action: str, field: str | None = None,
    old: str | None = None, new: str | None = None
) -> None:
    log = TaskAuditLog(
        task_id=task_id, user_id=user_id, action=action,
        field=field, old_value=old, new_value=new,
    )
    db.add(log)


async def _sync_directive_progress(db: AsyncSession, directive_id: int) -> None:
    """Recalculate directive.progress from tasks that have directive_id = directive_id."""
    from app.models.directive import Directive, DirectiveHistory

    directive = await db.get(Directive, directive_id)
    if not directive or directive.deleted_at is not None:
        return

    statuses = (await db.execute(
        select(Task.status).where(
            Task.directive_id == directive_id,
            Task.deleted_at.is_(None),
        )
    )).scalars().all()

    if not statuses:
        return

    done = sum(1 for s in statuses if s == "completed")
    new_prog = int(done / len(statuses) * 100)

    if new_prog == directive.progress:
        return

    old_prog = directive.progress
    directive.progress = new_prog

    if new_prog >= 100 and directive.status == "active":
        directive.status = "completed"
        db.add(DirectiveHistory(
            directive_id=directive_id,
            user_id=directive.issuer_id,
            action="auto_completed",
            old_status="active",
            new_status="completed",
            old_progress=old_prog,
            new_progress=new_prog,
        ))


async def _sync_program_progress(db: AsyncSession, program_id: int) -> None:
    """Recalculate program.progress_percent from tasks that have program_id = program_id."""
    from app.models.program import Program

    program = await db.get(Program, program_id)
    if not program or program.deleted_at is not None:
        return

    statuses = (await db.execute(
        select(Task.status).where(
            Task.program_id == program_id,
            Task.deleted_at.is_(None),
        )
    )).scalars().all()

    if not statuses:
        return

    done = sum(1 for s in statuses if s == "completed")
    new_prog = int(done / len(statuses) * 100)

    if new_prog == program.progress_percent:
        return

    program.progress_percent = new_prog


async def _sync_project_progress(db: AsyncSession, task_id: int) -> None:
    """Recalculate progress_percent for all strategic projects linked to this task."""
    from app.models.strategic import ProjectTaskLink, StrategicProject

    project_ids = (await db.execute(
        select(ProjectTaskLink.project_id).where(ProjectTaskLink.task_id == task_id)
    )).scalars().all()

    for project_id in project_ids:
        project = await db.get(StrategicProject, project_id)
        if not project:
            continue

        progresses = (await db.execute(
            select(Task.progress_percent).join(
                ProjectTaskLink, ProjectTaskLink.task_id == Task.id
            ).where(
                ProjectTaskLink.project_id == project_id,
                Task.deleted_at.is_(None),
            )
        )).scalars().all()

        if not progresses:
            continue

        new_prog = int(sum(progresses) / len(progresses))
        if new_prog != project.progress_percent:
            project.progress_percent = new_prog


async def _set_departments(
    db: AsyncSession, task: Task,
    lead_dept_id: int | None,
    coordinating_ids: list[int],
) -> None:
    await db.execute(
        delete(TaskDepartment).where(TaskDepartment.task_id == task.id)
    )
    await db.flush()

    if lead_dept_id:
        db.add(TaskDepartment(task_id=task.id, department_id=lead_dept_id, role="lead"))
    for dept_id in coordinating_ids:
        if dept_id != lead_dept_id:
            db.add(TaskDepartment(task_id=task.id, department_id=dept_id, role="coordinating"))


def _calc_stats(rows: list[Task]) -> TaskStats:
    now = datetime.now(timezone.utc)
    s = TaskStats(total=len(rows))
    progresses = []
    for t in rows:
        st = t.status
        if st == "pending":
            s.pending += 1
        elif st == "in_progress":
            s.in_progress += 1
        elif st == "completed":
            s.completed += 1
        elif st == "cancelled":
            s.cancelled += 1
        elif st == "overdue":
            s.overdue += 1
        # Real-time: pending/in_progress tasks past due_date (before scheduler ran)
        if st in ("pending", "in_progress") and _is_overdue(t):
            s.overdue += 1
        if t.priority == "high":
            s.high_priority += 1
        elif t.priority == "urgent":
            s.urgent_priority += 1
        progresses.append(t.progress_percent)
    s.avg_progress = round(sum(progresses) / len(progresses), 1) if progresses else 0.0
    return s


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/export/excel")
async def export_tasks_excel(
    status: str | None = None,
    priority: str | None = None,
    lead_dept_id: int | None = None,
    assignee_id: int | None = None,
    program_id: int | None = None,
    overdue_only: bool = False,
    due_before: datetime | None = None,
    due_after: datetime | None = None,
    search: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import io as _io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = select(Task).where(Task.deleted_at.is_(None))

    if current_user.role not in ("admin", "leader"):
        user_dept_id = await _get_user_dept_id(db, current_user.id)
        dept_task_ids = select(TaskDepartment.task_id).where(
            TaskDepartment.department_id == user_dept_id,
        ) if user_dept_id else select(TaskDepartment.task_id).where(False)
        q = q.where(
            or_(
                Task.lead_department_id == user_dept_id,
                Task.id.in_(dept_task_ids),
                Task.assignee_id == current_user.id,
                Task.created_by == current_user.id,
            )
        )

    if status == "overdue":
        _now = datetime.now(timezone.utc)
        q = q.where(Task.due_date.isnot(None), Task.due_date < _now, Task.status.notin_(["completed", "cancelled"]))
    elif status:
        q = q.where(Task.status == status)
    if priority:
        q = q.where(Task.priority == priority)
    if lead_dept_id:
        q = q.where(Task.lead_department_id == lead_dept_id)
    if assignee_id:
        q = q.where(Task.assignee_id == assignee_id)
    if program_id:
        q = q.where(Task.program_id == program_id)
    if due_before:
        q = q.where(Task.due_date <= due_before)
    if due_after:
        q = q.where(Task.due_date >= due_after)
    if search:
        term = f"%{search}%"
        q = q.where(or_(Task.title.ilike(term), Task.task_code.ilike(term)))
    if overdue_only:
        now = datetime.now(timezone.utc)
        q = q.where(Task.due_date < now, Task.status.notin_(["completed", "cancelled"]))

    q = q.order_by(Task.created_at.desc()).limit(5000)
    for ld in _LIST_LOADS:
        q = q.options(ld)
    rows = (await db.execute(q)).scalars().all()

    from app.models.program import Program as _Program
    prog_ids = {t.program_id for t in rows if t.program_id}
    prog_map: dict[int, str] = {}
    if prog_ids:
        progs = (await db.execute(select(_Program).where(_Program.id.in_(prog_ids)))).scalars().all()
        prog_map = {p.id: p.name or p.code or "" for p in progs}

    STATUS_LBL  = {"pending": "Chờ xử lý", "in_progress": "Đang thực hiện", "completed": "Hoàn thành", "cancelled": "Đã huỷ"}
    PRIORITY_LBL = {"low": "Thấp", "medium": "TB", "high": "Cao", "urgent": "Khẩn"}

    def fmt_dt(v):
        if not v: return ""
        try: return v.strftime("%d/%m/%Y")
        except: return str(v)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách nhiệm vụ"

    hdr_fill = PatternFill("solid", fgColor="1565C0")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="E3F2FD")

    COLS = ["Mã NV", "Tiêu đề", "Trạng thái", "Ưu tiên", "Tiến độ (%)", "Hạn xử lý", "Đơn vị CT", "Người thực hiện", "Chương trình", "Ngày tạo"]
    WIDTHS = [14, 48, 16, 10, 12, 14, 20, 22, 22, 14]

    for ci, (col, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell = ws.cell(1, ci, col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, t in enumerate(rows, 2):
        fill = alt_fill if ri % 2 == 0 else None
        assignee_name = ""
        if t.assignee_staff:
            assignee_name = t.assignee_staff.full_name or ""
        elif t.assignee:
            assignee_name = t.assignee.full_name or t.assignee.username or ""

        prog_name = prog_map.get(t.program_id, "") if t.program_id else ""

        vals = [
            t.task_code or "",
            t.title or "",
            STATUS_LBL.get(t.status, t.status or ""),
            PRIORITY_LBL.get(t.priority, t.priority or ""),
            t.progress_percent or 0,
            fmt_dt(t.due_date),
            t.lead_department.name if t.lead_department else "",
            assignee_name,
            prog_name,
            fmt_dt(t.created_at),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            if fill: cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import datetime as _dt
    fname = f"nhiem-vu-{_dt.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("", response_model=PaginatedTasks)
async def list_tasks(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    status: str | None = None,
    priority: str | None = None,
    lead_dept_id: int | None = None,
    assignee_id: int | None = None,
    assignee_staff_id: int | None = None,
    supervising_user_id: int | None = None,
    incoming_document_id: int | None = None,
    outgoing_document_id: int | None = None,
    directive_id: int | None = None,
    program_id: int | None = None,
    parent_task_id: int | None = None,
    task_type: str | None = None,
    is_project: bool | None = None,
    overdue_only: bool = False,
    due_before: datetime | None = None,
    due_after: datetime | None = None,
    coordinating_dept_id: int | None = None,
    search: str | None = None,
    sort_by: str = "created_at",
    sort_dir: str = "desc",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = select(Task).where(Task.deleted_at.is_(None))

    # ── Role-based visibility filter ──────────────────────────────────────────
    # admin + leader: see everything
    # manager + staff: see only tasks where their dept is lead/coordinating,
    #                  OR they are personally assigned
    if current_user.role not in ("admin", "leader"):
        user_dept_id = await _get_user_dept_id(db, current_user.id)
        dept_task_ids = select(TaskDepartment.task_id).where(
            TaskDepartment.department_id == user_dept_id,
        ) if user_dept_id else select(TaskDepartment.task_id).where(False)

        q = q.where(
            or_(
                Task.lead_department_id == user_dept_id,
                Task.id.in_(dept_task_ids),
                Task.assignee_id == current_user.id,
                Task.created_by == current_user.id,
            )
        )
    # ─────────────────────────────────────────────────────────────────────────

    if status == "overdue":
        # Virtual status: tasks past due_date that are not completed/cancelled
        _now = datetime.now(timezone.utc)
        q = q.where(
            Task.due_date.isnot(None),
            Task.due_date < _now,
            Task.status.notin_(["completed", "cancelled"]),
        )
    elif status:
        q = q.where(Task.status == status)
    if priority:
        q = q.where(Task.priority == priority)
    if lead_dept_id:
        q = q.where(Task.lead_department_id == lead_dept_id)
    if assignee_id:
        q = q.where(Task.assignee_id == assignee_id)
    if assignee_staff_id:
        q = q.where(Task.assignee_staff_id == assignee_staff_id)
    if supervising_user_id:
        q = q.where(Task.supervising_user_id == supervising_user_id)
    if incoming_document_id:
        q = q.where(Task.incoming_document_id == incoming_document_id)
    if outgoing_document_id:
        q = q.where(Task.outgoing_document_id == outgoing_document_id)
    if directive_id:
        q = q.where(Task.directive_id == directive_id)
    if program_id:
        q = q.where(Task.program_id == program_id)
    if parent_task_id:
        q = q.where(Task.parent_task_id == parent_task_id)
    if task_type:
        q = q.where(Task.task_type == task_type)
    if is_project is not None:
        q = q.where(Task.is_project == is_project)
    if due_before:
        q = q.where(Task.due_date <= due_before)
    if due_after:
        q = q.where(Task.due_date >= due_after)
    if search:
        term = f"%{search}%"
        q = q.where(or_(Task.title.ilike(term), Task.task_code.ilike(term), Task.content_summary.ilike(term)))
    if coordinating_dept_id:
        sub = select(TaskDepartment.task_id).where(
            TaskDepartment.department_id == coordinating_dept_id,
            TaskDepartment.role == "coordinating",
        )
        q = q.where(Task.id.in_(sub))
    if overdue_only:
        now = datetime.now(timezone.utc)
        q = q.where(Task.due_date < now, Task.status.notin_(["completed", "cancelled"]))

    sort_col = sort_by if sort_by in VALID_SORT_COLS else "created_at"
    col = getattr(Task, sort_col)
    q = q.order_by(col.desc() if sort_dir == "desc" else col.asc())

    total_result = await db.execute(select(func.count()).select_from(q.order_by(None).subquery()))
    total = total_result.scalar() or 0

    q = q.offset((page - 1) * page_size).limit(page_size)
    for ld in _LIST_LOADS:
        q = q.options(ld)

    rows = (await db.execute(q)).scalars().all()
    items = [TaskRead.model_validate(_to_read(t)) for t in rows]

    # Inject project_ids from ProjectTaskLink (single query, no N+1)
    if rows:
        from app.models.strategic import ProjectTaskLink
        task_ids = [t.id for t in rows]
        proj_links = (await db.execute(
            select(ProjectTaskLink.task_id, ProjectTaskLink.project_id)
            .where(ProjectTaskLink.task_id.in_(task_ids))
        )).all()
        proj_map: dict[int, list[int]] = {}
        for tid, pid in proj_links:
            proj_map.setdefault(tid, []).append(pid)
        for item in items:
            item.project_ids = proj_map.get(item.id, [])

    return PaginatedTasks(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        pages=ceil(total / page_size) if total else 1,
    )


@router.post("", response_model=TaskRead, status_code=201)
async def create_task(
    body: TaskCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if body.priority not in VALID_PRIORITIES:
        raise HTTPException(400, f"priority must be one of {VALID_PRIORITIES}")

    await _check_task_write_permission(db, current_user, body.lead_department_id, "create")

    # Validate FK references exist and are not soft-deleted
    if body.incoming_document_id:
        from app.models.document import Document
        doc = await db.get(Document, body.incoming_document_id)
        if not doc or getattr(doc, 'deleted_at', None):
            raise HTTPException(404, "Văn bản đến không tồn tại")
    if body.outgoing_document_id:
        from app.models.document import Document
        doc = await db.get(Document, body.outgoing_document_id)
        if not doc or getattr(doc, 'deleted_at', None):
            raise HTTPException(404, "Văn bản đi không tồn tại")
    if body.directive_id:
        from app.models.directive import Directive
        d = await db.get(Directive, body.directive_id)
        if not d or getattr(d, 'deleted_at', None):
            raise HTTPException(404, "Chỉ đạo không tồn tại")
    if body.program_id:
        from app.models.program import Program
        p = await db.get(Program, body.program_id)
        if not p or getattr(p, 'deleted_at', None):
            raise HTTPException(404, "Chương trình không tồn tại")

    if body.start_date and body.due_date and body.due_date < body.start_date:
        raise HTTPException(400, "Hạn hoàn thành không thể trước ngày bắt đầu")

    task_code = await _next_task_code(db)
    t = Task(
        task_code=task_code,
        title=body.title,
        description=body.description,
        content_summary=body.content_summary,
        priority=body.priority,
        is_project=body.is_project,
        project_type=body.project_type,
        budget_amount=body.budget_amount,
        budget_disbursed=body.budget_disbursed,
        status="pending",
        progress_percent=0,
        start_date=body.start_date,
        due_date=body.due_date,
        incoming_document_id=body.incoming_document_id,
        outgoing_document_id=body.outgoing_document_id,
        directive_id=body.directive_id,
        program_id=body.program_id,
        parent_task_id=body.parent_task_id,
        created_by=current_user.id,
        assignee_id=body.assignee_id,
        assignee_staff_id=body.assignee_staff_id,
        supervising_user_id=body.supervising_user_id,
        lead_department_id=body.lead_department_id,
        reminder_enabled=body.reminder_enabled,
    )
    db.add(t)
    await db.flush()

    await _set_departments(db, t, body.lead_department_id, body.coordinating_department_ids)
    await _audit(db, t.id, current_user.id, "created")
    await db.commit()

    task = await _get_task(db, t.id, detail=False)
    return TaskRead.model_validate(_to_read(task))


@router.get("/stats", response_model=TaskStats)
async def get_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    base_cond = [Task.deleted_at.is_(None)]

    # Same role-based filter as list_tasks
    if current_user.role not in ("admin", "leader"):
        user_dept_id = await _get_user_dept_id(db, current_user.id)
        dept_task_ids = select(TaskDepartment.task_id).where(
            TaskDepartment.department_id == user_dept_id,
        ) if user_dept_id else select(TaskDepartment.task_id).where(False)
        base_cond.append(
            or_(
                Task.lead_department_id == user_dept_id,
                Task.id.in_(dept_task_ids),
                Task.assignee_id == current_user.id,
                Task.created_by == current_user.id,
            )
        )

    row = (await db.execute(
        select(
            func.count().label("total"),
            func.sum(case((Task.status == "pending",     1), else_=0)).label("pending"),
            func.sum(case((Task.status == "in_progress", 1), else_=0)).label("in_progress"),
            func.sum(case((Task.status == "completed",   1), else_=0)).label("completed"),
            func.sum(case((Task.status == "cancelled",   1), else_=0)).label("cancelled"),
            func.sum(case((Task.priority == "high",      1), else_=0)).label("high_priority"),
            func.sum(case((Task.priority == "urgent",    1), else_=0)).label("urgent_priority"),
            func.avg(Task.progress_percent).label("avg_progress"),
        ).where(*base_cond)
    )).one()

    now = datetime.now(timezone.utc)
    overdue = (await db.execute(
        select(func.count()).where(
            *base_cond,
            Task.status.notin_(["completed", "cancelled"]),
            Task.due_date < now,
        )
    )).scalar_one()

    return TaskStats(
        total=row.total or 0,
        pending=row.pending or 0,
        in_progress=row.in_progress or 0,
        completed=row.completed or 0,
        cancelled=row.cancelled or 0,
        overdue=overdue,
        high_priority=row.high_priority or 0,
        urgent_priority=row.urgent_priority or 0,
        avg_progress=round(float(row.avg_progress or 0), 1),
    )


@router.get("/overdue", response_model=list[TaskRead])
async def get_overdue(
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    now = datetime.now(timezone.utc)
    q = (
        select(Task)
        .where(
            Task.deleted_at.is_(None),
            Task.due_date < now,
            Task.status.notin_(["completed", "cancelled"]),
        )
        .order_by(Task.due_date.asc())
        .limit(limit)
    )
    for ld in _LIST_LOADS:
        q = q.options(ld)
    rows = (await db.execute(q)).scalars().all()
    return [TaskRead.model_validate(_to_read(t)) for t in rows]


@router.get("/{task_id}", response_model=TaskReadDetail)
async def get_task(
    task_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    t = await _get_task(db, task_id, detail=True)
    d = _to_read(t)

    def _user_min(u: User | None) -> dict | None:
        if u is None:
            return None
        return {"id": u.id, "full_name": u.full_name, "username": u.username}

    def _doc_min(doc: Any) -> dict | None:
        if doc is None:
            return None
        return {"id": doc.id, "title": doc.title, "doc_number": getattr(doc, "doc_number", None)}

    def _dir_min(drv: Any) -> dict | None:
        if drv is None:
            return None
        return {"id": drv.id, "title": drv.title}

    d["updater"] = _user_min(t.updater)
    d["supervisor"] = _user_min(t.supervisor)
    d["incoming_document"] = _doc_min(t.incoming_document)
    d["outgoing_document"] = _doc_min(t.outgoing_document)
    d["directive"] = _dir_min(t.directive)
    d["departments"] = [
        {
            "id": td.id,
            "department_id": td.department_id,
            "role": td.role,
            "department": {"id": td.department.id, "name": td.department.name,
                           "short_name": getattr(td.department, "short_name", None),
                           "code": getattr(td.department, "code", None)} if td.department else None,
        }
        for td in t.departments
    ]
    d["comments"] = [
        {
            "id": c.id, "task_id": c.task_id, "user_id": c.user_id,
            "content": c.content, "created_at": c.created_at,
            "user": _user_min(c.user),
        }
        for c in t.comments
    ]
    d["attachments"] = [
        {
            "id": a.id, "task_id": a.task_id, "user_id": a.user_id,
            "filename": a.filename, "file_path": a.file_path,
            "file_size": a.file_size, "created_at": a.created_at,
        }
        for a in t.attachments
    ]
    d["audit_logs"] = [
        {
            "id": lg.id, "action": lg.action, "field": lg.field,
            "old_value": lg.old_value, "new_value": lg.new_value,
            "created_at": lg.created_at, "user": _user_min(lg.user),
        }
        for lg in t.audit_logs
    ]
    d["subtasks"] = [TaskRead.model_validate(_to_read(sub)) for sub in t.subtasks]
    return TaskReadDetail.model_validate(d)


@router.put("/{task_id}", response_model=TaskRead)
async def update_task(
    task_id: int,
    body: TaskUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    t = await _get_task(db, task_id)
    await _check_task_write_permission(db, current_user, t.lead_department_id, "update", task_id)

    if body.priority is not None and body.priority not in VALID_PRIORITIES:
        raise HTTPException(400, f"priority must be one of {VALID_PRIORITIES}")

    start = body.start_date if body.start_date is not None else t.start_date
    due = body.due_date if body.due_date is not None else t.due_date
    start_d = start.date() if hasattr(start, 'date') else start
    due_d = due.date() if hasattr(due, 'date') else due
    if start_d and due_d and due_d < start_d:
        raise HTTPException(400, "Hạn hoàn thành không thể trước ngày bắt đầu")

    if body.incoming_document_id:
        from app.models.document import Document as Doc
        doc = await db.get(Doc, body.incoming_document_id)
        if not doc or getattr(doc, 'deleted_at', None):
            raise HTTPException(404, "Văn bản đến không tồn tại")
    if body.outgoing_document_id:
        from app.models.document import Document as Doc
        doc = await db.get(Doc, body.outgoing_document_id)
        if not doc or getattr(doc, 'deleted_at', None):
            raise HTTPException(404, "Văn bản đi không tồn tại")
    if body.directive_id:
        from app.models.directive import Directive
        d = await db.get(Directive, body.directive_id)
        if not d or getattr(d, 'deleted_at', None):
            raise HTTPException(404, "Chỉ đạo không tồn tại")
    if body.program_id:
        from app.models.program import Program
        p = await db.get(Program, body.program_id)
        if not p or getattr(p, 'deleted_at', None):
            raise HTTPException(404, "Chương trình không tồn tại")

    old_program_id = t.program_id

    fields = ["title", "description", "content_summary", "priority", "is_project",
              "project_type", "budget_amount", "budget_disbursed",
              "start_date", "due_date",
              "incoming_document_id", "outgoing_document_id", "directive_id", "program_id",
              "assignee_id", "assignee_staff_id", "supervising_user_id", "lead_department_id",
              "reminder_enabled", "completion_note"]

    for f in fields:
        val = getattr(body, f)
        if val is not None:
            old = str(getattr(t, f))
            setattr(t, f, val)
            await _audit(db, t.id, current_user.id, "updated", f, old, str(val))

    t.updated_by = current_user.id

    if body.coordinating_department_ids is not None:
        await _set_departments(db, t, t.lead_department_id, body.coordinating_department_ids)

    # Sync program progress nếu program_id thay đổi
    new_program_id = t.program_id
    if old_program_id and old_program_id != new_program_id:
        await _sync_program_progress(db, old_program_id)
    if new_program_id:
        await _sync_program_progress(db, new_program_id)

    await db.commit()
    task = await _get_task(db, t.id, detail=False)
    return TaskRead.model_validate(_to_read(task))


@router.patch("/{task_id}/status", response_model=TaskRead)
async def update_status(
    task_id: int,
    body: TaskStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if body.status not in VALID_STATUSES:
        raise HTTPException(400, f"status must be one of {VALID_STATUSES}")

    t = await _get_task(db, task_id)
    old_status = t.status
    t.status = body.status
    t.updated_by = current_user.id

    if body.status == "completed":
        t.progress_percent = 100
        t.completed_at = datetime.now(timezone.utc)
        if body.completion_note:
            t.completion_note = body.completion_note
    else:
        if old_status == "completed":
            t.completed_at = None
        if body.status == "in_progress" and t.progress_percent == 0:
            t.progress_percent = 10

    await _audit(db, t.id, current_user.id, "status_changed", "status", old_status, body.status)

    if t.directive_id:
        await _sync_directive_progress(db, t.directive_id)
    if t.program_id:
        await _sync_program_progress(db, t.program_id)
    await _sync_project_progress(db, t.id)

    await db.commit()

    task = await _get_task(db, t.id, detail=False)
    return TaskRead.model_validate(_to_read(task))


@router.patch("/{task_id}/progress", response_model=TaskRead)
async def update_progress(
    task_id: int,
    body: TaskProgressUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not (0 <= body.progress_percent <= 100):
        raise HTTPException(400, "progress_percent must be 0-100")

    t = await _get_task(db, task_id)
    old_pct = t.progress_percent
    t.progress_percent = body.progress_percent
    t.updated_by = current_user.id

    if body.progress_percent == 100 and t.status not in ("completed", "cancelled"):
        t.status = "completed"
        t.completed_at = datetime.now(timezone.utc)
        if body.completion_note:
            t.completion_note = body.completion_note
    elif body.progress_percent > 0 and t.status == "pending":
        t.status = "in_progress"

    await _audit(db, t.id, current_user.id, "progress_updated", "progress_percent", str(old_pct), str(body.progress_percent))

    if t.directive_id:
        await _sync_directive_progress(db, t.directive_id)
    if t.program_id:
        await _sync_program_progress(db, t.program_id)
    await _sync_project_progress(db, t.id)

    await db.commit()

    task = await _get_task(db, t.id, detail=False)
    return TaskRead.model_validate(_to_read(task))


@router.delete("/{task_id}", status_code=204)
async def delete_task(
    task_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    t = await _get_task(db, task_id)
    await _check_task_write_permission(db, current_user, t.lead_department_id, "delete", task_id)
    t.deleted_at = datetime.now(timezone.utc)
    t.updated_by = current_user.id
    await _audit(db, t.id, current_user.id, "deleted")
    await db.commit()


# ── Comments ──────────────────────────────────────────────────────────────────

@router.post("/{task_id}/comments", response_model=TaskCommentRead, status_code=201)
async def add_comment(
    task_id: int,
    body: TaskCommentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_task(db, task_id)
    c = TaskComment(task_id=task_id, user_id=current_user.id, content=body.content)
    db.add(c)
    await db.flush()
    await _audit(db, task_id, current_user.id, "comment_added")
    await db.commit()

    q = select(TaskComment).where(TaskComment.id == c.id).options(selectinload(TaskComment.user))
    c = (await db.execute(q)).scalar_one()
    return TaskCommentRead.model_validate(c)


@router.delete("/{task_id}/comments/{comment_id}", status_code=204)
async def delete_comment(
    task_id: int,
    comment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(TaskComment).where(TaskComment.id == comment_id, TaskComment.task_id == task_id)
    )
    c = result.scalar_one_or_none()
    if c is None:
        raise HTTPException(404, "Comment not found")
    if c.user_id != current_user.id and current_user.role not in ("admin", "manager"):
        raise HTTPException(403, "Forbidden")
    await db.delete(c)
    await db.commit()


# ── Attachments ───────────────────────────────────────────────────────────────

@router.post("/{task_id}/attachments", response_model=TaskAttachmentRead, status_code=201)
async def upload_attachment(
    task_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_task(db, task_id)
    # Strip directory traversal — use only the bare filename, prefix with uuid
    safe_name = Path(file.filename or "file").name
    ext = Path(safe_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTS:
        raise HTTPException(
            400,
            f"Định dạng file không được phép: '{ext or 'không rõ'}'. "
            f"Chấp nhận: {', '.join(sorted(ALLOWED_UPLOAD_EXTS))}"
        )
    dest = UPLOAD_DIR / f"{uuid.uuid4().hex}_{safe_name}"
    with dest.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    size = dest.stat().st_size

    a = TaskAttachment(
        task_id=task_id, user_id=current_user.id,
        filename=safe_name,
        file_path=str(dest), file_size=size,
    )
    db.add(a)
    await db.flush()
    await _audit(db, task_id, current_user.id, "attachment_added", "filename", None, file.filename)
    await db.commit()
    await db.refresh(a)
    return TaskAttachmentRead.model_validate(a)


@router.delete("/{task_id}/attachments/{attachment_id}", status_code=204)
async def delete_attachment(
    task_id: int,
    attachment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(TaskAttachment).where(TaskAttachment.id == attachment_id, TaskAttachment.task_id == task_id)
    )
    a = result.scalar_one_or_none()
    if a is None:
        raise HTTPException(404, "Attachment not found")
    try:
        Path(a.file_path).unlink(missing_ok=True)
    except Exception:
        pass
    await db.delete(a)
    await db.commit()


# ── Departments ───────────────────────────────────────────────────────────────

@router.post("/{task_id}/departments", response_model=TaskDeptRead, status_code=201)
async def add_department(
    task_id: int,
    body: TaskDeptAdd,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_task(db, task_id)
    existing = (await db.execute(
        select(TaskDepartment).where(
            TaskDepartment.task_id == task_id,
            TaskDepartment.department_id == body.department_id,
        )
    )).scalar_one_or_none()
    if existing:
        existing.role = body.role
        await db.commit()
        await db.refresh(existing)
        return TaskDeptRead.model_validate(existing)

    td = TaskDepartment(task_id=task_id, department_id=body.department_id, role=body.role)
    db.add(td)
    await db.commit()
    await db.refresh(td)
    return TaskDeptRead.model_validate(td)


@router.delete("/{task_id}/departments/{dept_id}", status_code=204)
async def remove_department(
    task_id: int,
    dept_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(TaskDepartment).where(
            TaskDepartment.task_id == task_id,
            TaskDepartment.department_id == dept_id,
        )
    )
    td = result.scalar_one_or_none()
    if td is None:
        raise HTTPException(404, "Department link not found")
    await db.delete(td)
    await db.commit()


# ── Excel Import ──────────────────────────────────────────────────────────────

@router.get("/import/template")
async def download_task_template(_: User = Depends(get_current_user)):
    from fastapi.responses import Response
    from app.services.excel_import import task_template
    data = task_template()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mau_import_nhiem_vu.xlsx"},
    )


@router.post("/import")
async def import_tasks_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.services.excel_import import parse_tasks
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Chỉ chấp nhận file .xlsx hoặc .xls")

    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(400, "File không được vượt quá 5 MB")

    records, errors = parse_tasks(data)
    if not records and errors:
        raise HTTPException(422, {"errors": errors})

    VALID = {"low", "medium", "high", "urgent"}
    imported = 0
    for r in records:
        due_dt = None
        if r.get("due_date_str"):
            try:
                due_dt = datetime.strptime(r["due_date_str"], "%d/%m/%Y").replace(
                    hour=17, minute=0, tzinfo=timezone.utc
                )
            except ValueError:
                pass
        start_dt = None
        if r.get("start_date_str"):
            try:
                start_dt = datetime.strptime(r["start_date_str"], "%d/%m/%Y").date()
            except ValueError:
                pass
        priority = r["priority"] if r["priority"] in VALID else "medium"
        task_code = await _next_task_code(db)
        t = Task(
            task_code=task_code,
            title=r["title"],
            description=r["description"] or None,
            content_summary=r["responsible_unit"] or None,
            priority=priority,
            status="pending",
            progress_percent=0,
            start_date=start_dt,
            due_date=due_dt,
            created_by=current_user.id,
        )
        db.add(t)
        await db.flush()
        await _audit(db, t.id, current_user.id, "created")
        imported += 1

    await db.commit()
    return {"imported": imported, "errors": errors}
