import logging
import shutil
from datetime import datetime, timezone

logger = logging.getLogger(__name__)
from math import ceil
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from fastapi.responses import FileResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import get_current_user, require_admin_or_leader
from app.models.department import Department
from app.models.document import Document, DocumentComment, DocumentHistory, DocumentTask
from app.models.staff import Staff
from app.models.task import Task, TaskAuditLog
from app.models.user import User
from app.schemas.document import (
    DocumentCommentCreate,
    DocumentCommentRead,
    DocumentCreate,
    DocumentHistoryRead,
    DocumentRead,
    DocumentReadDetail,
    DocumentStatusUpdate,
    DocumentTaskCreate,
    DocumentTaskRead,
    DocumentUpdate,
    PaginatedResponse,
)

from app.services import ai_parser_service, ocr_service
from app.core.config import settings as _settings
router = APIRouter()
DOC_UPLOAD_DIR = Path(_settings.UPLOAD_DIR) / "documents"
DOC_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

VALID_STATUSES = {"pending", "processing", "done", "archived"}
VALID_TYPES = {"incoming", "outgoing", "internal"}
VALID_PRIORITIES = {"normal", "urgent", "very_urgent"}
ALLOWED_UPLOAD_EXTS = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".png", ".jpg", ".jpeg"}


# ─── Helpers ─────────────────────────────────────────────────────────────────

async def _get_doc_or_404(db: AsyncSession, doc_id: int) -> Document:
    result = await db.execute(
        select(Document).where(Document.id == doc_id, Document.deleted_at.is_(None))
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    return doc


async def _doc_with_relations(db: AsyncSession, doc_id: int) -> Document:
    stmt = (
        select(Document)
        .options(
            selectinload(Document.creator),
            selectinload(Document.assignee),
            selectinload(Document.responsible_department),
            selectinload(Document.assignee_staff),
        )
        .where(Document.id == doc_id, Document.deleted_at.is_(None))
    )
    return (await db.execute(stmt)).scalar_one()


async def _doc_detail(db: AsyncSession, doc_id: int) -> Document:
    stmt = (
        select(Document)
        .options(
            selectinload(Document.creator),
            selectinload(Document.assignee),
            selectinload(Document.responsible_department),
            selectinload(Document.assignee_staff),
            selectinload(Document.comments).selectinload(DocumentComment.user),
            selectinload(Document.history).selectinload(DocumentHistory.user),
            selectinload(Document.linked_tasks)
            .selectinload(DocumentTask.task)
            .selectinload(Task.assignee),
        )
        .where(Document.id == doc_id, Document.deleted_at.is_(None))
    )
    doc = (await db.execute(stmt)).scalar_one_or_none()
    if not doc:
        raise HTTPException(status_code=404, detail="Không tìm thấy văn bản")
    return doc


def _add_history(
    db: AsyncSession,
    doc_id: int,
    user_id: int,
    action: str,
    old_status: str | None = None,
    new_status: str | None = None,
    note: str | None = None,
) -> None:
    db.add(DocumentHistory(
        doc_id=doc_id,
        user_id=user_id,
        action=action,
        old_status=old_status,
        new_status=new_status,
        note=note,
    ))


# ─── AI Parse ────────────────────────────────────────────────────────────────

_AI_PARSE_ALLOWED = {".pdf", ".jpg", ".jpeg", ".png", ".docx", ".doc", ".txt"}


@router.post("/ai-parse")
async def ai_parse_document(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """OCR + AI-parse a file; return extracted fields to auto-fill the form.
    No DB record is created.
    """
    import asyncio as _asyncio

    ext = Path(file.filename or "file").suffix.lower()
    if ext not in _AI_PARSE_ALLOWED:
        raise HTTPException(400, f"Chỉ hỗ trợ PDF, JPG, PNG. Nhận được: {ext or 'không rõ'}")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(400, "File quá lớn (tối đa 20 MB)")

    tmp_path = ocr_service.save_upload(content, file.filename or "upload.bin")
    try:
        # Vision-first: Gemini reads the document as images (PDF scan + digital both work)
        result = await _asyncio.to_thread(ai_parser_service.parse_file_with_vision, tmp_path)
    finally:
        ocr_service.delete_file(str(tmp_path))

    vb = result.get("van_ban", {})

    summary_points: list[str] = vb.get("summary_points") or []
    keywords: list[str] = vb.get("tu_khoa") or []

    summary: str | None = None
    if summary_points:
        summary = "\n".join(f"• {p}" for p in summary_points)

    return {
        "doc_number": vb.get("so_ky_hieu"),
        "title": vb.get("trich_yeu"),
        "issuer": vb.get("co_quan_ban_hanh"),
        "category": vb.get("loai_van_ban"),
        "issue_date": vb.get("ngay_ban_hanh"),
        "summary": summary,
        "summary_points": summary_points or None,
        "keywords": keywords or None,
    }


# ─── Upload & AI Analyze → Save ──────────────────────────────────────────────

@router.post("/upload", response_model=DocumentRead, status_code=status.HTTP_201_CREATED)
async def upload_and_analyze(
    file: UploadFile = File(...),
    doc_type: str = "incoming",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload file → OCR → Gemini AI analyze → Save to DB → Return document.

    Supports: PDF, DOCX, DOC, TXT, JPG, PNG.
    Scans AI result into: title, issuer, category, issue_date, summary, keywords, domain, raw_text.
    """
    import asyncio as _asyncio
    import shutil as _shutil

    ext = Path(file.filename or "file").suffix.lower()
    if ext not in _AI_PARSE_ALLOWED:
        raise HTTPException(
            400,
            f"Không hỗ trợ định dạng {ext or 'không rõ'}. Chấp nhận: PDF, DOCX, TXT, JPG, PNG.",
        )

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        raise HTTPException(400, "File quá lớn (tối đa 50 MB)")

    tmp_path = ocr_service.save_upload(content, file.filename or "upload.bin")
    try:
        result = await _asyncio.to_thread(ai_parser_service.parse_file_with_vision, tmp_path)
    except Exception as exc:
        ocr_service.delete_file(str(tmp_path))
        raise HTTPException(500, f"Lỗi phân tích AI: {exc}")

    vb = result.get("van_ban", {})
    summary_points: list[str] = vb.get("summary_points") or []
    keywords_list: list[str] = vb.get("tu_khoa") or []
    domain: str | None = vb.get("linh_vuc")

    # Build summary text from bullet points
    summary_text: str | None = None
    if summary_points:
        summary_text = "\n".join(f"• {p}" for p in summary_points)

    # Parse issue_date
    from datetime import date as _date
    issue_date: _date | None = None
    raw_date = vb.get("ngay_ban_hanh")
    if raw_date:
        try:
            issue_date = _date.fromisoformat(str(raw_date))
        except Exception:
            pass

    # Title fallback: use first summary point or filename
    title = (
        vb.get("so_ky_hieu")
        or (summary_points[0][:200] if summary_points else None)
        or Path(file.filename or "").stem
        or "Văn bản chưa đặt tên"
    )

    # Raw text for later re-analysis
    raw_text_for_db: str | None = None
    if ext not in {".pdf", ".jpg", ".jpeg", ".png"}:
        # For text-based files we can get raw text cheaply
        try:
            raw_text_for_db, _ = await _asyncio.to_thread(ocr_service.ocr_file, tmp_path)
        except Exception:
            pass

    # Create Document record
    doc = Document(
        doc_number=vb.get("so_ky_hieu") or None,
        title=title,
        doc_type=doc_type,
        category=vb.get("loai_van_ban") or None,
        issuer=vb.get("co_quan_ban_hanh") or None,
        issue_date=issue_date,
        status="pending",
        priority="normal",
        summary=summary_text,
        keywords=keywords_list,
        domain=domain,
        raw_text=raw_text_for_db,
        ai_processed=True,
        created_by=current_user.id,
    )
    db.add(doc)
    await db.flush()
    _add_history(db, doc.id, current_user.id, "uploaded_ai")
    await db.commit()

    # Persist file to permanent storage
    import uuid as _uuid_mod
    dest_dir = DOC_UPLOAD_DIR / str(doc.id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    stored_name = f"{_uuid_mod.uuid4().hex}{ext}"
    dest = dest_dir / stored_name
    try:
        _shutil.copy2(str(tmp_path), str(dest))
        doc.file_name = file.filename or stored_name  # original name shown to user
        doc.file_path = str(dest.relative_to(DOC_UPLOAD_DIR.parent))
        doc.file_size = len(content)
        doc.file_mime = file.content_type or "application/octet-stream"
        await db.commit()
    except Exception as exc:
        logger.warning("Could not move uploaded file: %s", exc)
    finally:
        ocr_service.delete_file(str(tmp_path))

    return await _doc_with_relations(db, doc.id)


# ─── List / Create ────────────────────────────────────────────────────────────

@router.get("", response_model=PaginatedResponse[DocumentRead])
async def list_documents(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=500),
    search: str | None = Query(None),
    doc_type: str | None = Query(None),
    doc_status: str | None = Query(None, alias="status"),
    priority: str | None = Query(None),
    issuer: str | None = Query(None),
    assignee_id: int | None = Query(None),
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    conditions = [Document.deleted_at.is_(None)]
    if search:
        conditions.append(or_(
            Document.title.ilike(f"%{search}%"),
            Document.doc_number.ilike(f"%{search}%"),
            Document.issuer.ilike(f"%{search}%"),
        ))
    if doc_type:
        conditions.append(Document.doc_type == doc_type)
    if doc_status:
        conditions.append(Document.status == doc_status)
    if priority:
        conditions.append(Document.priority == priority)
    if issuer:
        conditions.append(Document.issuer.ilike(f"%{issuer}%"))
    if assignee_id:
        conditions.append(Document.assignee_id == assignee_id)
    if from_date:
        from datetime import date
        try:
            conditions.append(Document.issue_date >= date.fromisoformat(from_date))
        except ValueError:
            pass
    if to_date:
        from datetime import date
        try:
            conditions.append(Document.issue_date <= date.fromisoformat(to_date))
        except ValueError:
            pass

    count_stmt = select(func.count(Document.id)).where(*conditions)
    total = (await db.execute(count_stmt)).scalar_one()

    stmt = (
        select(Document)
        .options(
            selectinload(Document.creator),
            selectinload(Document.assignee),
            selectinload(Document.responsible_department),
            selectinload(Document.assignee_staff),
        )
        .where(*conditions)
        .order_by(Document.created_at.desc())
        .offset((page - 1) * size)
        .limit(size)
    )
    docs = (await db.execute(stmt)).scalars().all()

    return PaginatedResponse(
        items=list(docs),
        total=total,
        page=page,
        size=size,
        pages=max(1, ceil(total / size)),
    )


@router.get("/export")
async def export_documents(
    search: str | None = Query(None),
    doc_type: str | None = Query(None),
    doc_status: str | None = Query(None, alias="status"),
    priority: str | None = Query(None),
    from_date: str | None = Query(None),
    to_date: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export filtered documents to Excel (.xlsx)."""
    import io as _io
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Same filter logic as list_documents ──────────────────────────────────
    conditions = [Document.deleted_at.is_(None)]
    if search:
        conditions.append(or_(
            Document.title.ilike(f"%{search}%"),
            Document.doc_number.ilike(f"%{search}%"),
            Document.issuer.ilike(f"%{search}%"),
        ))
    if doc_type:
        conditions.append(Document.doc_type == doc_type)
    if doc_status:
        conditions.append(Document.status == doc_status)
    if priority:
        conditions.append(Document.priority == priority)
    if from_date:
        from datetime import date as _date
        try:
            conditions.append(Document.issue_date >= _date.fromisoformat(from_date))
        except ValueError:
            pass
    if to_date:
        from datetime import date as _date
        try:
            conditions.append(Document.issue_date <= _date.fromisoformat(to_date))
        except ValueError:
            pass

    stmt = (
        select(Document)
        .options(
            selectinload(Document.assignee),
            selectinload(Document.responsible_department),
            selectinload(Document.assignee_staff),
        )
        .where(*conditions)
        .order_by(Document.created_at.desc())
        .limit(5000)
    )
    docs = (await db.execute(stmt)).scalars().all()

    # ── Build Excel ───────────────────────────────────────────────────────────
    TYPE_LABELS  = {"incoming": "Văn bản đến", "outgoing": "Văn bản đi", "internal": "Nội bộ"}
    STATUS_LABELS = {"pending": "Chờ xử lý", "processing": "Đang xử lý", "done": "Đã xử lý", "archived": "Lưu trữ"}
    PRIORITY_LABELS = {"normal": "Thường", "urgent": "Khẩn", "very_urgent": "Hỏa tốc"}

    def fmt_date(val):
        if not val:
            return ""
        try:
            return val.strftime("%d/%m/%Y")
        except Exception:
            return str(val)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách văn bản"

    # Title row
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = "DANH SÁCH VĂN BẢN"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:N2")
    sub = ws["A2"]
    sub.value = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   Tổng số: {len(docs)} văn bản"
    sub.font = Font(italic=True, size=10)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Header row
    headers = [
        "STT", "Số hiệu", "Trích yếu", "Loại", "Hình thức",
        "Cơ quan ban hành", "Ngày ban hành", "Ngày nhận",
        "Hạn xử lý", "Trạng thái", "Độ ưu tiên",
        "Đơn vị thực hiện", "Người xử lý", "Ghi chú",
    ]
    col_widths = [5, 16, 42, 14, 14, 22, 14, 12, 12, 14, 11, 20, 18, 20]

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 32

    # Data rows
    even_fill = PatternFill("solid", fgColor="F0F4FA")
    data_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top")

    for row_idx, doc in enumerate(docs, start=4):
        is_even = (row_idx % 2 == 0)
        fill = even_fill if is_even else PatternFill()

        assignee_name = ""
        if doc.assignee_staff:
            assignee_name = doc.assignee_staff.full_name or ""
        elif doc.assignee:
            assignee_name = doc.assignee.full_name or doc.assignee.username or ""

        dept_name = doc.responsible_department.name if doc.responsible_department else ""

        row_data = [
            row_idx - 3,
            doc.doc_number or "",
            doc.title or "",
            TYPE_LABELS.get(doc.doc_type, doc.doc_type),
            doc.category or "",
            doc.issuer or "",
            fmt_date(doc.issue_date),
            fmt_date(doc.received_date),
            fmt_date(doc.deadline),
            STATUS_LABELS.get(doc.status, doc.status),
            PRIORITY_LABELS.get(doc.priority, doc.priority),
            dept_name,
            assignee_name,
            doc.summary or "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.fill = fill
            cell.alignment = center_align if col_idx in (1, 4, 5, 7, 8, 9, 10, 11) else data_align
        ws.row_dimensions[row_idx].height = 20

    # Freeze header rows
    ws.freeze_panes = "A4"

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"van-ban-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _can_manage_doc(user: User, doc: Document, user_dept_id: int | None = None) -> bool:
    if user.role in ('admin', 'leader'):
        return True
    if doc.created_by == user.id:
        return True
    if user.role == 'manager' and user_dept_id and doc.responsible_department_id == user_dept_id:
        return True
    return False


async def _get_user_dept(db: AsyncSession, user_id: int) -> int | None:
    from app.models.staff import Staff
    staff = (await db.execute(select(Staff).where(Staff.user_id == user_id))).scalar_one_or_none()
    return staff.department_id if staff else None


@router.post("", response_model=DocumentRead, status_code=status.HTTP_201_CREATED)
async def create_document(
    body: DocumentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    program_id = body.program_id
    doc = Document(**body.model_dump(exclude={"program_id"}), created_by=current_user.id)
    db.add(doc)
    await db.flush()
    if program_id:
        from app.models.program import DocumentProgram
        db.add(DocumentProgram(
            document_id=doc.id, program_id=program_id,
            link_type="implements", created_by=current_user.id,
        ))
    _add_history(db, doc.id, current_user.id, "created")
    await db.commit()
    return await _doc_with_relations(db, doc.id)


# ─── Detail / Update / Delete ────────────────────────────────────────────────

@router.get("/{doc_id}", response_model=DocumentReadDetail)
async def get_document(
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    return await _doc_detail(db, doc_id)


@router.put("/{doc_id}", response_model=DocumentRead)
async def update_document(
    doc_id: int,
    body: DocumentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = await _get_doc_or_404(db, doc_id)
    if not _can_manage_doc(current_user, doc, await _get_user_dept(db, current_user.id)):
        raise HTTPException(403, "Không có quyền sửa văn bản này")
    changes = body.model_dump(exclude_unset=True, exclude={"program_id"})
    for field, value in changes.items():
        setattr(doc, field, value)
    if "program_id" in body.model_fields_set and body.program_id is not None:
        from app.models.program import DocumentProgram
        existing = (await db.execute(
            select(DocumentProgram).where(DocumentProgram.document_id == doc_id)
        )).scalars().all()
        for lnk in existing:
            await db.delete(lnk)
        db.add(DocumentProgram(
            document_id=doc_id, program_id=body.program_id,
            link_type="implements", created_by=current_user.id,
        ))
    _add_history(db, doc.id, current_user.id, "updated")
    await db.commit()
    return await _doc_with_relations(db, doc_id)


@router.delete("/{doc_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_document(
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = await _get_doc_or_404(db, doc_id)
    if not _can_manage_doc(current_user, doc, await _get_user_dept(db, current_user.id)):
        raise HTTPException(403, "Không có quyền xóa văn bản này")
    doc.deleted_at = datetime.now(timezone.utc)
    _add_history(db, doc.id, current_user.id, "deleted")
    await db.commit()


@router.post("/{doc_id}/restore", response_model=DocumentRead)
async def restore_document(
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_leader),
):
    result = await db.execute(
        select(Document).where(Document.id == doc_id, Document.deleted_at.isnot(None))
    )
    doc = result.scalar_one_or_none()
    if not doc:
        raise HTTPException(404, "Không tìm thấy văn bản đã xóa")
    doc.deleted_at = None
    _add_history(db, doc.id, current_user.id, "restored")
    await db.commit()
    return await _doc_with_relations(db, doc_id)


# ─── Status ───────────────────────────────────────────────────────────────────

@router.patch("/{doc_id}/status", response_model=DocumentRead)
async def update_status(
    doc_id: int,
    body: DocumentStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin_or_leader),
):
    if body.status not in VALID_STATUSES:
        raise HTTPException(422, f"Trạng thái không hợp lệ: {body.status}")
    doc = await _get_doc_or_404(db, doc_id)
    old = doc.status
    doc.status = body.status
    _add_history(db, doc.id, current_user.id, "status_changed",
                 old_status=old, new_status=body.status, note=body.note)
    await db.commit()
    return await _doc_with_relations(db, doc_id)


# ─── File Upload / Serve ──────────────────────────────────────────────────────

@router.post("/{doc_id}/file", response_model=DocumentRead)
async def upload_file(
    doc_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = await _get_doc_or_404(db, doc_id)
    if not _can_manage_doc(current_user, doc, await _get_user_dept(db, current_user.id)):
        raise HTTPException(403, "Không có quyền upload file cho văn bản này")

    safe_name = Path(file.filename or "file").name
    ext = Path(safe_name).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTS:
        raise HTTPException(
            400,
            f"Định dạng file không được phép: '{ext or 'không rõ'}'. "
            f"Chấp nhận: {', '.join(sorted(ALLOWED_UPLOAD_EXTS))}"
        )

    doc_dir = DOC_UPLOAD_DIR / str(doc_id)
    doc_dir.mkdir(exist_ok=True)

    # remove old file
    if doc.file_path:
        old = Path(doc.file_path)
        if old.exists():
            old.unlink()
    import uuid as _uuid
    stored_name = f"{_uuid.uuid4().hex}{ext}"
    file_path = doc_dir / stored_name
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    doc.file_name = safe_name          # original name shown to user
    doc.file_path = str(file_path)     # UUID-based path on disk
    doc.file_size = file_path.stat().st_size
    doc.file_mime = file.content_type or "application/octet-stream"

    _add_history(db, doc.id, current_user.id, "file_uploaded", note=safe_name)
    await db.commit()
    return await _doc_with_relations(db, doc_id)


@router.get("/{doc_id}/file")
async def serve_file(
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    doc = await _get_doc_or_404(db, doc_id)
    if not doc.file_path:
        raise HTTPException(404, "Văn bản chưa có file đính kèm")
    path = Path(doc.file_path)
    if not path.exists():
        raise HTTPException(404, "File không tồn tại trên server")

    mime = doc.file_mime or "application/octet-stream"
    disposition = "inline" if mime == "application/pdf" else "attachment"
    return FileResponse(
        path=str(path),
        filename=doc.file_name,
        media_type=mime,
        headers={"Content-Disposition": f'{disposition}; filename="{doc.file_name}"'},
    )


# ─── Comments ─────────────────────────────────────────────────────────────────

@router.post("/{doc_id}/comments", response_model=DocumentCommentRead,
             status_code=status.HTTP_201_CREATED)
async def add_comment(
    doc_id: int,
    body: DocumentCommentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_doc_or_404(db, doc_id)
    comment = DocumentComment(doc_id=doc_id, user_id=current_user.id, content=body.content)
    db.add(comment)
    _add_history(db, doc_id, current_user.id, "commented")
    await db.flush()
    await db.commit()
    stmt = (
        select(DocumentComment)
        .options(selectinload(DocumentComment.user))
        .where(DocumentComment.id == comment.id)
    )
    return (await db.execute(stmt)).scalar_one()


@router.delete("/{doc_id}/comments/{comment_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_comment(
    doc_id: int,
    comment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(DocumentComment).where(
            DocumentComment.id == comment_id,
            DocumentComment.doc_id == doc_id,
        )
    )
    comment = result.scalar_one_or_none()
    if not comment:
        raise HTTPException(404, "Không tìm thấy bình luận")
    if comment.user_id != current_user.id:
        raise HTTPException(403, "Không có quyền xóa bình luận này")
    await db.delete(comment)
    await db.commit()


# ─── Tasks from Document ──────────────────────────────────────────────────────

@router.post("/{doc_id}/tasks", response_model=DocumentTaskRead,
             status_code=status.HTTP_201_CREATED)
async def create_task_from_doc(
    doc_id: int,
    body: DocumentTaskCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_doc_or_404(db, doc_id)

    from app.api.v1.endpoints.tasks import _next_task_code
    task_code = await _next_task_code(db)

    task = Task(
        task_code=task_code,
        title=body.title,
        description=body.description,
        priority=body.priority,
        due_date=body.deadline,
        assignee_id=body.assignee_id,
        lead_department_id=body.lead_department_id,
        created_by=current_user.id,
        status="pending",
        progress_percent=0,
    )
    db.add(task)
    await db.flush()

    db.add(TaskAuditLog(
        task_id=task.id,
        user_id=current_user.id,
        action="created",
        new_value=f"Từ văn bản #{doc_id}",
    ))

    link = DocumentTask(doc_id=doc_id, task_id=task.id)
    db.add(link)
    _add_history(db, doc_id, current_user.id, "task_created", note=body.title)

    await db.flush()
    await db.commit()

    stmt = (
        select(DocumentTask)
        .options(
            selectinload(DocumentTask.task).selectinload(Task.assignee)
        )
        .where(DocumentTask.id == link.id)
    )
    return (await db.execute(stmt)).scalar_one()


# ─── AI Extract Tasks + Bulk Create ─────────────────────────────────────────

class AISuggestedTask(BaseModel):
    title: str
    description: str | None = None
    deadline: str | None = None      # YYYY-MM-DD
    priority: str = "medium"
    lead_agency: str | None = None   # tên đơn vị dạng text từ AI


class AIExtractResponse(BaseModel):
    tasks: list[AISuggestedTask]
    source: str   # "ai_file" | "ai_text" | "none"


class BulkTaskItem(BaseModel):
    title: str
    description: str | None = None
    deadline: datetime | None = None
    priority: str = "medium"
    lead_department_id: int | None = None
    assignee_id: int | None = None


class BulkTaskRequest(BaseModel):
    tasks: list[BulkTaskItem]


class BulkTaskResponse(BaseModel):
    created: int
    task_ids: list[int]


@router.post("/{doc_id}/extract-tasks", response_model=AIExtractResponse)
async def extract_tasks_ai(
    doc_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Chạy AI trên file đã lưu của văn bản, trả về danh sách nhiệm vụ gợi ý."""
    import asyncio as _asyncio

    doc = await _get_doc_or_404(db, doc_id)

    result = None
    source = "none"

    # Ưu tiên file gốc → AI Vision
    if doc.file_path:
        file_path = DOC_UPLOAD_DIR.parent / doc.file_path
        if file_path.exists():
            try:
                result = await _asyncio.to_thread(ai_parser_service.parse_file_with_vision, file_path)
                source = "ai_file"
            except Exception as exc:
                logger.warning("extract_tasks_ai: vision failed for doc %s: %s", doc_id, exc)

    # Fallback: raw_text
    if not result and doc.raw_text:
        try:
            result = await _asyncio.to_thread(ai_parser_service.parse_document, doc.raw_text)
            source = "ai_text"
        except Exception as exc:
            logger.warning("extract_tasks_ai: text parse failed for doc %s: %s", doc_id, exc)

    if not result:
        return AIExtractResponse(tasks=[], source="none")

    tasks: list[AISuggestedTask] = []
    for t in result.get("nhiem_vu") or []:
        raw_dl = t.get("deadline")
        deadline_str: str | None = None
        if raw_dl:
            try:
                from datetime import date as _date
                deadline_str = str(_date.fromisoformat(str(raw_dl)))
            except Exception:
                pass

        prio_map = {"urgent": "urgent", "high": "high", "medium": "medium", "low": "low"}
        tasks.append(AISuggestedTask(
            title=(t.get("ten_nhiem_vu") or "")[:300].strip(),
            description=t.get("mo_ta") or None,
            deadline=deadline_str,
            priority=prio_map.get(t.get("muc_uu_tien", "medium"), "medium"),
            lead_agency=t.get("don_vi_chu_tri") or None,
        ))

    return AIExtractResponse(tasks=[t for t in tasks if t.title], source=source)


@router.post("/{doc_id}/bulk-tasks", response_model=BulkTaskResponse, status_code=201)
async def bulk_create_tasks(
    doc_id: int,
    body: BulkTaskRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Tạo nhiều nhiệm vụ cùng lúc từ danh sách gợi ý AI, liên kết với văn bản."""
    await _get_doc_or_404(db, doc_id)

    if not body.tasks:
        raise HTTPException(400, "Danh sách nhiệm vụ trống")

    from app.api.v1.endpoints.tasks import _next_task_code

    created_ids: list[int] = []
    for item in body.tasks:
        task_code = await _next_task_code(db)
        task = Task(
            task_code=task_code,
            title=item.title,
            description=item.description,
            priority=item.priority if item.priority in ("low", "medium", "high", "urgent") else "medium",
            due_date=item.deadline,
            assignee_id=item.assignee_id,
            lead_department_id=item.lead_department_id,
            created_by=current_user.id,
            status="pending",
            progress_percent=0,
        )
        db.add(task)
        await db.flush()

        db.add(TaskAuditLog(
            task_id=task.id,
            user_id=current_user.id,
            action="created",
            new_value=f"AI từ văn bản #{doc_id}",
        ))
        db.add(DocumentTask(doc_id=doc_id, task_id=task.id))
        _add_history(db, doc_id, current_user.id, "task_created", note=item.title)
        created_ids.append(task.id)

    await db.commit()
    return BulkTaskResponse(created=len(created_ids), task_ids=created_ids)


# ─── Capture từ dhtn.dcs.vn (Bookmarklet) ───────────────────────────────────

class CaptureRequest(BaseModel):
    """Minimal data sent by the bookmarklet from dhtn.dcs.vn."""
    title: str
    doc_number: str | None = None
    doc_type: str = "incoming"      # incoming | outgoing
    issuer: str | None = None       # Đơn vị ban hành / nơi gửi
    trich_yeu: str | None = None    # Trích yếu nội dung
    issue_date: str | None = None   # dd/mm/yyyy from dhtn
    source_url: str | None = None   # URL trang dhtn.dcs.vn
    do_mat: str | None = None       # Độ mật (Thường / Mật / ...)
    # Auto-create task?
    create_task: bool = False
    task_title: str | None = None
    task_due_date: str | None = None
    lead_department_id: int | None = None


class CaptureResponse(BaseModel):
    doc_id: int
    doc_number: str | None
    task_id: int | None = None
    message: str


@router.post("/capture", response_model=CaptureResponse, status_code=201)
async def capture_from_dhtn(
    body: CaptureRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Receive a document captured from dhtn.dcs.vn via bookmarklet.
    Optionally auto-create a linked task.
    """
    from datetime import date as dt_date
    import re

    # Check for duplicate doc_number before creating
    if body.doc_number and body.doc_number.strip():
        existing = (await db.execute(
            select(Document.id, Document.title)
            .where(Document.doc_number == body.doc_number.strip(), Document.deleted_at.is_(None))
        )).first()
        if existing:
            raise HTTPException(409, detail={
                "code": "DUPLICATE_DOC_NUMBER",
                "message": f"Văn bản ký hiệu '{body.doc_number}' đã tồn tại trong hệ thống.",
                "existing_doc_id": existing.id,
                "existing_doc_title": existing.title or body.doc_number,
            })

    # Parse issue_date (dd/mm/yyyy)
    issue_date = None
    if body.issue_date:
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                issue_date = datetime.strptime(body.issue_date, fmt).date()
                break
            except ValueError:
                pass

    doc = Document(
        doc_number=body.doc_number,
        title=body.title,
        doc_type=body.doc_type,
        issuer=body.issuer,
        summary=body.trich_yeu,          # trích yếu → summary
        category=body.do_mat or None,    # độ mật → category
        issue_date=issue_date,
        received_date=dt_date.today() if body.doc_type == "incoming" else None,
        status="pending",
        priority="normal",
        created_by=current_user.id,
    )
    db.add(doc)
    await db.flush()

    _add_history(db, doc.id, current_user.id, "created",
                 note=f"Nhập từ dhtn.dcs.vn — {body.source_url or ''}")

    task_id = None
    if body.create_task and body.task_title:
        from app.models.task import Task as TaskModel, TaskDepartment as TDept
        from app.api.v1.endpoints.tasks import _next_task_code
        task_code = await _next_task_code(db)
        due_dt = None
        if body.task_due_date:
            try:
                due_dt = datetime.strptime(body.task_due_date, "%d/%m/%Y")
            except ValueError:
                pass
        t = TaskModel(
            task_code=task_code,
            title=body.task_title or body.title,
            incoming_document_id=doc.id if body.doc_type == "incoming" else None,
            outgoing_document_id=doc.id if body.doc_type == "outgoing" else None,
            lead_department_id=body.lead_department_id,
            due_date=due_dt,
            status="pending",
            priority="medium",
            created_by=current_user.id,
        )
        db.add(t)
        await db.flush()
        db.add(DocumentTask(doc_id=doc.id, task_id=t.id))
        if body.lead_department_id:
            db.add(TDept(task_id=t.id, department_id=body.lead_department_id, role="lead"))
        task_id = t.id

    await db.commit()

    return CaptureResponse(
        doc_id=doc.id,
        doc_number=doc.doc_number,
        task_id=task_id,
        message=f"Đã nhập văn bản{' và tạo nhiệm vụ' if task_id else ''} thành công",
    )
