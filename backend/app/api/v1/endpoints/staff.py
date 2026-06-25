"""Staff management endpoint.

Staff is the primary identity record for all personnel.
Creating a Staff member with email + password auto-creates (or links) a User account
so all FK → users.id references continue to work transparently.
"""
from __future__ import annotations

from math import ceil
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel, EmailStr
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import get_current_user, require_admin_or_leader
from app.core.security import hash_password, verify_password
from app.models.department import Department
from app.models.staff import Staff
from app.models.user import User

router = APIRouter()


def _like(s: str) -> str:
    return s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


VALID_ROLES = {"admin", "leader", "manager", "staff"}
ROLE_LABELS = {
    "admin":   "Admin",
    "leader":  "Lãnh đạo",
    "manager": "Quản lý",
    "staff":   "Nhân viên",
}

# ─── Schemas ──────────────────────────────────────────────────────────────────

class DeptMin(BaseModel):
    model_config = {"from_attributes": True}
    id: int
    name: str
    short_name: str | None = None
    code: str | None = None


class UserMin(BaseModel):
    model_config = {"from_attributes": True}
    id: int
    username: str
    full_name: str | None = None


class StaffRead(BaseModel):
    model_config = {"from_attributes": True}
    id: int
    employee_code: str | None = None
    full_name: str
    position: str | None = None
    phone: str | None = None
    email: str | None = None
    avatar_url: str | None = None
    note: str | None = None
    role: str
    is_active: bool
    has_password: bool = False   # True nếu có thể đăng nhập
    department_id: int | None = None
    user_id: int | None = None
    department: DeptMin | None = None
    user: UserMin | None = None


class StaffCreate(BaseModel):
    full_name: str
    email: EmailStr | None = None
    password: str | None = None          # if provided → auto-create User + auth
    role: str = "staff"
    employee_code: str | None = None
    position: str | None = None
    department_id: int | None = None
    phone: str | None = None
    avatar_url: str | None = None
    note: str | None = None
    is_active: bool = True


class StaffUpdate(BaseModel):
    full_name: str | None = None
    email: EmailStr | None = None
    role: str | None = None
    position: str | None = None
    department_id: int | None = None
    phone: str | None = None
    avatar_url: str | None = None
    note: str | None = None
    is_active: bool | None = None
    employee_code: str | None = None


class StaffPasswordReset(BaseModel):
    new_password: str


class PaginatedStaff(BaseModel):
    items: list[StaffRead]
    total: int
    page: int
    size: int
    pages: int


# ─── Helpers ──────────────────────────────────────────────────────────────────

async def _get_or_404(db: AsyncSession, staff_id: int) -> Staff:
    stmt = (
        select(Staff)
        .options(selectinload(Staff.department), selectinload(Staff.user))
        .where(Staff.id == staff_id, Staff.deleted_at.is_(None))
    )
    s = (await db.execute(stmt)).scalar_one_or_none()
    if not s:
        raise HTTPException(404, "Không tìm thấy nhân sự")
    return s


async def _ensure_user(db: AsyncSession, staff: Staff) -> User:
    """
    Ensure the staff record has a linked User account.
    Creates one if missing. Keeps User.role in sync with Staff.role.
    Returns the User.
    """
    if staff.user_id:
        user = await db.get(User, staff.user_id)
        if user:
            if user.role != staff.role:
                user.role = staff.role
            if user.is_active != staff.is_active:
                user.is_active = staff.is_active
            return user

    # No linked user → create one
    email = staff.email or f"staff_{staff.id}@system.local"
    username = (email.split("@")[0] + str(staff.id))[:50]

    # Check username uniqueness
    existing = (await db.execute(select(User).where(User.username == username))).scalar_one_or_none()
    if existing:
        username = f"{username}_{staff.id}"

    user = User(
        username=username,
        email=email,
        hashed_password=staff.password_hash or hash_password(__import__('secrets').token_hex(16)),
        full_name=staff.full_name,
        role=staff.role,
        is_active=staff.is_active,
    )
    db.add(user)
    await db.flush()   # get user.id
    staff.user_id = user.id
    return user


# ─── Routes ───────────────────────────────────────────────────────────────────

@router.get("", response_model=PaginatedStaff)
async def list_staff(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=500),
    search: str | None = Query(None),
    department_id: int | None = Query(None),
    role: str | None = Query(None),
    active_only: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    conditions = [Staff.deleted_at.is_(None)]
    if search:
        conditions.append(or_(
            Staff.full_name.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.position.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.employee_code.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.phone.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.email.ilike(f"%{_like(search)}%", escape="\\"),
        ))
    if department_id:
        conditions.append(Staff.department_id == department_id)
    if role:
        conditions.append(Staff.role == role)
    if active_only:
        conditions.append(Staff.is_active.is_(True))

    base_q = select(Staff).where(*conditions) if conditions else select(Staff)
    total = (await db.execute(select(func.count()).select_from(base_q.subquery()))).scalar_one()

    stmt = (
        base_q
        .options(selectinload(Staff.department), selectinload(Staff.user))
        .order_by(Staff.full_name)
        .offset((page - 1) * size).limit(size)
    )
    items = (await db.execute(stmt)).scalars().all()
    return PaginatedStaff(
        items=items, total=total, page=page, size=size,
        pages=max(1, ceil(total / size)),
    )


@router.get("/export/excel")
async def export_staff_excel(
    search: str | None = Query(None),
    department_id: int | None = Query(None),
    role: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    from datetime import datetime as _dt

    conditions = [Staff.deleted_at.is_(None)]
    if search:
        conditions.append(or_(
            Staff.full_name.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.position.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.employee_code.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.phone.ilike(f"%{_like(search)}%", escape="\\"),
            Staff.email.ilike(f"%{_like(search)}%", escape="\\"),
        ))
    if department_id:
        conditions.append(Staff.department_id == department_id)
    if role:
        conditions.append(Staff.role == role)

    base_q = select(Staff).where(*conditions)
    stmt = base_q.options(selectinload(Staff.department), selectinload(Staff.user)).order_by(Staff.full_name)
    items = (await db.execute(stmt)).scalars().all()

    ROLE_MAP = {"admin": "Admin", "leader": "Lãnh đạo", "manager": "Quản lý", "staff": "Nhân viên"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nhân sự"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1D4ED8")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["STT", "Mã NV", "Họ và tên", "Chức vụ", "Đơn vị", "Phân quyền",
               "Email", "Điện thoại", "Tên đăng nhập", "Trạng thái"]
    col_widths = [6, 10, 25, 22, 22, 14, 28, 15, 18, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill(fill_type="solid", fgColor="EFF6FF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for i, s in enumerate(items, 1):
        row = i + 1
        fill = alt_fill if i % 2 == 0 else None
        def c(col, val, align=left):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = align
            if fill: cell.fill = fill
        c(1, i, center)
        c(2, s.employee_code or "")
        c(3, s.full_name)
        c(4, s.position or "")
        c(5, (s.department.short_name or s.department.name) if s.department else "")
        c(6, ROLE_MAP.get(s.role, s.role), center)
        c(7, s.email or "")
        c(8, s.phone or "")
        c(9, s.user.username if s.user else "")
        c(10, "Hoạt động" if s.is_active else "Ngừng hoạt động", center)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"nhan-su-{_dt.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/dropdown")
async def staff_dropdown(
    search: str | None = Query(None),
    department_id: int | None = Query(None),
    role: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Lightweight endpoint for dropdowns — returns minimal fields, active staff only."""
    conditions = [Staff.is_active.is_(True)]
    if search:
        conditions.append(or_(
            Staff.full_name.ilike(f"%{search}%"),
            Staff.employee_code.ilike(f"%{search}%"),
            Staff.position.ilike(f"%{search}%"),
        ))
    if department_id:
        conditions.append(Staff.department_id == department_id)
    if role:
        conditions.append(Staff.role == role)

    stmt = (
        select(
            Staff.id,
            Staff.full_name,
            Staff.employee_code,
            Staff.position,
            Staff.department_id,
            Staff.role,
            Staff.email,
            Staff.user_id,
        )
        .where(*conditions)
        .order_by(Staff.full_name)
        .limit(200)
    )
    rows = (await db.execute(stmt)).all()

    # Attach department short_name
    dept_ids = {r.department_id for r in rows if r.department_id}
    depts: dict[int, str] = {}
    if dept_ids:
        dept_rows = (await db.execute(
            select(Department.id, Department.short_name).where(Department.id.in_(dept_ids))
        )).all()
        depts = {r.id: (r.short_name or "") for r in dept_rows}

    return [
        {
            "id":            r.id,
            "full_name":     r.full_name,
            "employee_code": r.employee_code,
            "position":      r.position,
            "department_id": r.department_id,
            "dept_name":     depts.get(r.department_id, ""),
            "role":          r.role,
            "role_label":    ROLE_LABELS.get(r.role, r.role),
            "email":         r.email,
            "user_id":       r.user_id,
        }
        for r in rows
    ]


@router.get("/roles")
async def get_role_options(_: User = Depends(get_current_user)):
    """Return available roles for dropdowns."""
    return [{"value": k, "label": v} for k, v in ROLE_LABELS.items()]


@router.post("", response_model=StaffRead, status_code=status.HTTP_201_CREATED)
async def create_staff(
    body: StaffCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader", "manager"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Không có quyền thêm nhân sự")

    # Manager: chỉ thêm được nhân viên (role=staff) cho đơn vị của mình
    if current_user.role == "manager":
        mgr_staff = (await db.execute(
            select(Staff).where(Staff.user_id == current_user.id)
        )).scalar_one_or_none()
        mgr_dept_id = mgr_staff.department_id if mgr_staff else None
        if not mgr_dept_id:
            raise HTTPException(403, "Không tìm thấy đơn vị của bạn")
        if body.department_id != mgr_dept_id:
            raise HTTPException(403, "Quản lý chỉ được thêm nhân viên cho đơn vị mình")
        if body.role not in ("staff",):
            raise HTTPException(403, "Quản lý chỉ được thêm nhân viên với quyền 'Nhân viên'")

    if body.role not in VALID_ROLES:
        raise HTTPException(400, f"Role không hợp lệ. Chọn: {list(VALID_ROLES)}")

    # Auto employee_code — use max existing number to avoid duplicates after deletions
    employee_code = body.employee_code
    if not employee_code:
        from sqlalchemy import cast, Integer
        result = await db.execute(
            select(func.max(
                cast(func.substr(Staff.employee_code, 3), Integer)
            )).where(Staff.employee_code.op("~")(r"^NS\d+$"))
        )
        max_num = result.scalar_one_or_none() or 0
        employee_code = f"NS{max_num + 1:03d}"

    pwd_hash = hash_password(body.password) if body.password else None

    s = Staff(
        full_name=body.full_name,
        email=str(body.email) if body.email else None,
        position=body.position,
        department_id=body.department_id,
        phone=body.phone,
        avatar_url=body.avatar_url,
        note=body.note,
        employee_code=employee_code,
        role=body.role,
        password_hash=pwd_hash,
        is_active=body.is_active,
    )
    db.add(s)
    await db.flush()

    # Auto-create linked User account (required for FK integrity across all modules)
    await _ensure_user(db, s)

    await db.commit()
    return await _get_or_404(db, s.id)


@router.get("/{staff_id}", response_model=StaffRead)
async def get_staff(
    staff_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    return await _get_or_404(db, staff_id)


@router.put("/{staff_id}", response_model=StaffRead)
async def update_staff(
    staff_id: int,
    body: StaffUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "leader"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Chỉ Admin/Lãnh đạo mới được sửa nhân sự")

    s = await _get_or_404(db, staff_id)

    if body.role is not None and body.role not in VALID_ROLES:
        raise HTTPException(400, f"Role không hợp lệ. Chọn: {list(VALID_ROLES)}")

    updates = body.model_dump(exclude_none=True)
    for k, v in updates.items():
        if k == "email":
            setattr(s, k, str(v))
        else:
            setattr(s, k, v)

    # Sync role + is_active to linked User
    if s.user_id:
        user = await db.get(User, s.user_id)
        if user:
            if body.role is not None:
                user.role = s.role
            if body.is_active is not None:
                user.is_active = s.is_active

    s.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return await _get_or_404(db, staff_id)


@router.post("/{staff_id}/reset-password", status_code=status.HTTP_204_NO_CONTENT)
async def reset_password(
    staff_id: int,
    body: StaffPasswordReset,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Admin/leader can reset any staff password. Staff can reset their own."""
    s = await _get_or_404(db, staff_id)

    is_self = s.user_id is not None and s.user_id == current_user.id
    if not is_self and current_user.role not in ("admin", "leader"):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Không có quyền đổi mật khẩu này")

    if len(body.new_password) < 8:
        raise HTTPException(400, "Mật khẩu phải có ít nhất 8 ký tự")

    new_hash = hash_password(body.new_password)
    s.password_hash = new_hash
    s.updated_at = datetime.now(timezone.utc)

    # Sync to linked User
    if s.user_id:
        user = await db.get(User, s.user_id)
        if user:
            user.hashed_password = new_hash

    await db.commit()


@router.delete("/{staff_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_staff(
    staff_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != "admin":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Chỉ Admin mới được xóa nhân sự")
    s = await _get_or_404(db, staff_id)
    s.deleted_at = datetime.now(timezone.utc)
    s.is_active = False
    await db.commit()
