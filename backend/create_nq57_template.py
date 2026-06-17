"""
Tạo file mẫu Excel NQ57 và seed thẳng vào database.

Chạy trên VPS:
  docker compose exec backend python /app/create_nq57_template.py
"""
import asyncio
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select


def create_excel(out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BC NQ57"

    for col, w in zip("ABCDEF", [6, 44, 26, 12, 14, 32]):
        ws.column_dimensions[col].width = w

    thin = Side(style="thin")
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def S(r, c, value="", bold=False, italic=False, size=11, ha="left",
          fill=None, wrap=False, border=False):
        cl = ws.cell(row=r, column=c, value=value)
        cl.font = Font(name="Times New Roman", bold=bold, italic=italic, size=size)
        cl.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
        if fill:
            cl.fill = PatternFill("solid", fgColor=fill)
        if border:
            cl.border = full_border
        return cl

    def merge(r, c1, c2):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)

    for r, h in [(1,24),(2,18),(3,18),(4,10),(5,20),(6,20),(7,10),
                 (8,38),(9,5),(10,46),(11,5),(12,10),(13,20),(14,55)]:
        ws.row_dimensions[r].height = h

    # Tiêu đề
    S(1,1,"BÁO CÁO TIẾN ĐỘ THEO NGHỊ QUYẾT 57",bold=True,size=14,ha="center")
    merge(1,1,6)
    S(2,1,"({{ky_bao_cao}} đến ngày {{den_ngay}})",italic=True,size=11,ha="center")
    merge(2,1,6)
    S(3,1,"Đơn vị: {{ten_don_vi}}",size=11)
    merge(3,1,6)

    # Tổng quan
    S(5,1,"TỔNG QUAN KỲ BÁO CÁO",bold=True,size=11,fill="D9E1F2",border=True)
    merge(5,1,6)
    for c in range(2,7):
        ws.cell(5,c).fill = PatternFill("solid",fgColor="D9E1F2")
        ws.cell(5,c).border = full_border
    for c,v,bold in [(1,"Tổng nhiệm vụ NQ57:",True),(2,"{{tong_nq57}}",False),
                     (3,"Đã hoàn thành:",True),(4,"{{nq57_hoan_thanh}}",False),
                     (5,"Tiến độ TB:",True),(6,"{{ti_le_nq57}}",False)]:
        S(6,c,v,bold=bold,size=10,border=True)

    # Header bảng
    for c,h in enumerate(["STT","Nội dung nhiệm vụ","Đơn vị thực hiện",
                           "Tiến độ","Hạn hoàn thành","Kết quả / Khó khăn"],1):
        S(8,c,h,bold=True,size=10,ha="center",fill="BDD7EE",wrap=True,border=True)

    # Vòng lặp
    ws.cell(9,1,"{{#danh_sach_nq57}}").font = Font(name="Times New Roman",size=7,color="AAAAAA")
    for c,v in enumerate(["{{item.stt}}","{{item.ten}}","{{item.don_vi}}",
                           "{{item.tien_do}}","{{item.han}}",""],1):
        ha = "center" if c in (1,4,5) else "left"
        S(10,c,v,size=10,ha=ha,wrap=True,border=True)
    ws.cell(11,1,"{{/danh_sach_nq57}}").font = Font(name="Times New Roman",size=7,color="AAAAAA")

    # Chữ ký
    S(13,1,"Ngày lập báo cáo: {{ngay_bao_cao}}",italic=True,size=10)
    merge(13,1,3)
    S(13,5,"THỦ TRƯỞNG ĐƠN VỊ",bold=True,size=10,ha="center")
    merge(13,5,6)
    S(14,5,"(Ký, đóng dấu)",italic=True,size=10,ha="center")
    merge(14,5,6)

    # Sheet hướng dẫn biến
    ws2 = wb.create_sheet("Hướng dẫn biến")
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 42
    for i,(a,b) in enumerate([
        ("BIẾN SCALAR",""),
        ("{{ky_bao_cao}}","Nhãn kỳ báo cáo, vd: Tháng 05/2026"),
        ("{{den_ngay}}","Ngày kết thúc kỳ"),
        ("{{tu_ngay}}","Ngày bắt đầu kỳ"),
        ("{{ten_don_vi}}","Tên đơn vị"),
        ("{{tong_nq57}}","Tổng số nhiệm vụ NQ57"),
        ("{{nq57_hoan_thanh}}","Số nhiệm vụ hoàn thành"),
        ("{{ti_le_nq57}}","Tiến độ NQ57 (%)"),
        ("{{ngay_bao_cao}}","Ngày in báo cáo"),
        ("",""),
        ("BIẾN VÒNG LẶP {{#danh_sach_nq57}}",""),
        ("{{item.stt}}","Số thứ tự"),
        ("{{item.ma}}","Mã nhiệm vụ"),
        ("{{item.ten}}","Nội dung nhiệm vụ"),
        ("{{item.nhom}}","Nhóm nhiệm vụ"),
        ("{{item.don_vi}}","Đơn vị thực hiện"),
        ("{{item.tien_do}}","Tiến độ (%)"),
        ("{{item.trang_thai}}","Trạng thái"),
        ("{{item.han}}","Hạn hoàn thành"),
        ("",""),
        ("LƯU Ý:","Cột 'Kết quả / Khó khăn' điền tay sau khi xuất"),
    ],1):
        ws2.cell(i,1,a).font = Font(name="Times New Roman",
                                     bold=bool(a) and "{{" not in a, size=10)
        ws2.cell(i,2,b).font = Font(name="Times New Roman",size=10)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"  File: {out_path}  ({out_path.stat().st_size:,} bytes)")


async def seed_db(file_path: Path) -> None:
    from app.core.database import AsyncSessionLocal
    from app.core.config import settings
    from app.models.report_template import ReportTemplate
    from app.models.user import User
    from app.services import template_engine

    async with AsyncSessionLocal() as db:
        # Lấy admin đầu tiên làm created_by
        admin = (await db.execute(
            select(User).where(User.role == "admin").limit(1)
        )).scalar_one_or_none()
        if not admin:
            admin = (await db.execute(select(User).limit(1))).scalar_one_or_none()
        if not admin:
            print("  ⚠️  Không tìm thấy user nào, bỏ qua seed DB")
            return

        name = "Báo cáo tiến độ NQ57 tuần"
        category = "bao_cao_thang"

        existing = (await db.execute(
            select(ReportTemplate).where(
                ReportTemplate.name == name,
                ReportTemplate.category == category,
            )
        )).scalar_one_or_none()

        if existing:
            print(f"  ⚠️  Đã tồn tại (id={existing.id}), bỏ qua")
            return

        # Copy file vào thư mục templates
        dest_dir = Path(settings.UPLOAD_DIR) / "templates" / category
        dest_dir.mkdir(parents=True, exist_ok=True)
        safe_name = re.sub(r"[^\w\-.]", "_", name)
        dest = dest_dir / f"{safe_name}_v1.xlsx"
        dest.write_bytes(file_path.read_bytes())

        scalars, lists = await asyncio.to_thread(
            template_engine.parse_variables, str(dest)
        )

        tpl = ReportTemplate(
            name=name,
            category=category,
            description="Báo cáo tiến độ thực hiện Nghị quyết 57 hàng tuần — tự động điền số liệu từ hệ thống",
            file_ext="xlsx",
            file_path=str(dest),
            file_size=dest.stat().st_size,
            variables_json=scalars,
            list_variables_json=lists,
            version=1,
            is_active=True,
            created_by=admin.id,
        )
        db.add(tpl)
        await db.commit()
        await db.refresh(tpl)
        print(f"  ✅ Đã thêm: id={tpl.id}  '{tpl.name}'  category={tpl.category}")
        print(f"     Scalar vars : {scalars}")
        print(f"     List vars   : {lists}")


async def main():
    tmp = Path("/tmp/mau_nq57.xlsx")
    print("── Bước 1: Tạo file Excel ──")
    create_excel(tmp)
    print("── Bước 2: Seed vào database ──")
    await seed_db(tmp)
    print("── Hoàn thành ──")

asyncio.run(main())
