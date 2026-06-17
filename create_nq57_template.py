"""
Chạy trên VPS:
  docker compose exec backend python /app/create_nq57_template.py
File được lưu tại: /app/uploads/reports/exports/mau_bao_cao_nq57.xlsx
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BC NQ57"

# Column widths: STT | Nội dung | Đơn vị | Tiến độ | Hạn | Kết quả/Khó khăn
for col, w in zip("ABCDEF", [6, 44, 26, 12, 14, 32]):
    ws.column_dimensions[col].width = w

thin = Side(style="thin")
full_border = Border(left=thin, right=thin, top=thin, bottom=thin)


def S(r, c, value="", bold=False, italic=False, size=11, ha="left",
      fill=None, wrap=False, border=False):
    cl = ws.cell(row=r, column=c, value=value)
    cl.font = Font(name="Times New Roman", bold=bold, italic=italic, size=size)
    cl.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fill:
        cl.fill = PatternFill("solid", fgColor=fill)
    if border:
        cl.border = full_border
    return cl


def merge(r, c1, c2):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)


# Row heights
for r, h in [(1, 24), (2, 18), (3, 18), (4, 10),
             (5, 20), (6, 20), (7, 10),
             (8, 38), (9, 5), (10, 46), (11, 5),
             (12, 10), (13, 20), (14, 55)]:
    ws.row_dimensions[r].height = h

# ── TIÊU ĐỀ ──────────────────────────────────────────────────────────────────
S(1, 1, "BÁO CÁO TIẾN ĐỘ THEO NGHỊ QUYẾT 57",
  bold=True, size=14, ha="center")
merge(1, 1, 6)

S(2, 1, "({{ky_bao_cao}} đến ngày {{den_ngay}})",
  italic=True, size=11, ha="center")
merge(2, 1, 6)

S(3, 1, "Đơn vị: {{ten_don_vi}}", size=11)
merge(3, 1, 6)

# ── TỔNG QUAN ────────────────────────────────────────────────────────────────
S(5, 1, "TỔNG QUAN KỲ BÁO CÁO", bold=True, size=11,
  fill="D9E1F2", border=True)
for c in range(2, 7):
    ws.cell(5, c).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(5, c).border = full_border
merge(5, 1, 6)

summary = [
    (1, "Tổng nhiệm vụ NQ57:", True),
    (2, "{{tong_nq57}}", False),
    (3, "Đã hoàn thành:", True),
    (4, "{{nq57_hoan_thanh}}", False),
    (5, "Tiến độ TB:", True),
    (6, "{{ti_le_nq57}}", False),
]
for c, v, bold in summary:
    S(6, c, v, bold=bold, size=10, border=True)

# ── BẢNG TIÊU ĐỀ CỘT ────────────────────────────────────────────────────────
headers = [
    "STT", "Nội dung nhiệm vụ", "Đơn vị thực hiện",
    "Tiến độ", "Hạn hoàn thành", "Kết quả / Khó khăn",
]
for c, h in enumerate(headers, 1):
    S(8, c, h, bold=True, size=10, ha="center",
      fill="BDD7EE", wrap=True, border=True)

# ── VÒNG LẶP NQ57 ────────────────────────────────────────────────────────────
# Marker dòng bắt đầu — chỉ ô A, text nhỏ màu xám (template engine đọc)
ws.cell(9, 1, "{{#danh_sach_nq57}}")
ws.cell(9, 1).font = Font(name="Times New Roman", size=7, color="AAAAAA")

# Dòng mẫu — mỗi NQ57 task sẽ sinh 1 dòng như này
tpl_row = [
    "{{item.stt}}",
    "{{item.ten}}",
    "{{item.don_vi}}",
    "{{item.tien_do}}",
    "{{item.han}}",
    "",   # Kết quả / Khó khăn — điền tay
]
for c, v in enumerate(tpl_row, 1):
    ha = "center" if c in (1, 4, 5) else "left"
    S(10, c, v, size=10, ha=ha, wrap=True, border=True)

# Marker kết thúc
ws.cell(11, 1, "{{/danh_sach_nq57}}")
ws.cell(11, 1).font = Font(name="Times New Roman", size=7, color="AAAAAA")

# ── CHỮ KÝ ───────────────────────────────────────────────────────────────────
S(13, 1, "Ngày lập báo cáo: {{ngay_bao_cao}}", italic=True, size=10)
merge(13, 1, 3)
S(13, 5, "THỦ TRƯỞNG ĐƠN VỊ", bold=True, size=10, ha="center")
merge(13, 5, 6)
S(14, 5, "(Ký, đóng dấu)", italic=True, size=10, ha="center")
merge(14, 5, 6)

# ── GHI CHÚ hướng dẫn (sheet 2) ─────────────────────────────────────────────
ws2 = wb.create_sheet("Hướng dẫn")
notes = [
    ("Biến tự động điền", ""),
    ("{{ky_bao_cao}}", "Nhãn kỳ báo cáo, vd: Tháng 05/2026"),
    ("{{den_ngay}}", "Ngày kết thúc kỳ, vd: 31/05/2026"),
    ("{{tu_ngay}}", "Ngày bắt đầu kỳ"),
    ("{{ten_don_vi}}", "Tên đơn vị (cấu hình hệ thống)"),
    ("{{tong_nq57}}", "Tổng số nhiệm vụ NQ57"),
    ("{{nq57_hoan_thanh}}", "Số nhiệm vụ NQ57 hoàn thành"),
    ("{{ti_le_nq57}}", "Tiến độ NQ57 (%)"),
    ("{{ngay_bao_cao}}", "Ngày in báo cáo (hôm nay)"),
    ("", ""),
    ("Biến trong vòng lặp {{#danh_sach_nq57}}", ""),
    ("{{item.stt}}", "Số thứ tự"),
    ("{{item.ma}}", "Mã NQ57"),
    ("{{item.ten}}", "Nội dung nhiệm vụ"),
    ("{{item.nhom}}", "Nhóm nhiệm vụ"),
    ("{{item.don_vi}}", "Đơn vị thực hiện"),
    ("{{item.tien_do}}", "Tiến độ (%)"),
    ("{{item.trang_thai}}", "Trạng thái"),
    ("{{item.han}}", "Hạn hoàn thành"),
    ("", ""),
    ("LƯU Ý:", "Cột 'Kết quả / Khó khăn' cần điền tay sau khi xuất"),
]
ws2.column_dimensions["A"].width = 38
ws2.column_dimensions["B"].width = 40
for i, (a, b) in enumerate(notes, 1):
    ws2.cell(i, 1, a).font = Font(name="Times New Roman",
                                   bold=("{{" not in a and a != ""),
                                   size=10, color="CC0000" if a.startswith("LƯU") else "000000")
    ws2.cell(i, 2, b).font = Font(name="Times New Roman", size=10)

# ── LƯU ─────────────────────────────────────────────────────────────────────
out = Path("/app/uploads/reports/exports/mau_bao_cao_nq57.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out))
print(f"✅ Đã tạo: {out}")
